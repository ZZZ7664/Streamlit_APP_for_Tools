# - * - 编码：UTF-8 - * -

# ================================================================================
# 统一导入区域 (最终版 v6)
# ================================================================================
import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import base64
import os
import plotly.graph_objects as go
import zipfile
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.styles.colors import Color
import traceback
import math
from pathlib import Path
from collections import defaultdict
import shutil
import multiprocessing

ROOT_DIR = Path(__file__).parent          # 项目根目录
ASSETS_DIR = ROOT_DIR / "assets"          # assets 文件夹

# ================================================================================
# ================================================================================
#  SCRIPT 1: 核心常量与规则库
# ================================================================================
# ================================================================================

# ★ 1. 测试项目规则库
ALL_TESTS = {
    'BMI': {'col_grade': '体重指数(BMI)_BMI', 'col_score': '体重指数(BMI)_分数', 'name': 'BMI指数'},
    '肺活量': {'col_grade': '肺活量_成绩', 'col_score': '肺活量_分数', 'name': '肺活量'},
    '50米跑': {'col_grade': '50米跑_成绩', 'col_score': '50米跑_分数', 'name': '50米跑'},
    '坐位体前屈': {'col_grade': '坐位体前屈_成绩', 'col_score': '坐位体前屈_分数', 'name': '坐位体前屈'},
    '立定跳远': {'col_grade': '立定跳远_成绩', 'col_score': '立定跳远_分数', 'name': '立定跳远'},
    '1分钟跳绳': {'col_grade': '一分钟跳绳_成绩', 'col_score': '一分钟跳绳_分数', 'name': '1分钟跳绳'},
    '1分钟仰卧起坐': {'col_grade': '一分钟仰卧起坐_成绩', 'col_score': '一分钟仰卧起坐_分数', 'name': '1分钟仰卧起坐'},
    '50米×8往返跑': {'col_grade': '50米×8往返跑_成绩', 'col_score': '50米×8往返跑_分数', 'name': '50米×8往返跑'},
    '引体向上': {'col_grade': '引体向上_成绩', 'col_score': '引体向上_分数', 'name': '引体向上'},
    '仰卧起坐(女)': {'col_grade': '一分钟仰卧起坐_成绩', 'col_score': '一分钟仰卧起坐_分数', 'name': '仰卧起坐'},
    '1000米跑': {'col_grade': '1000米跑(男)/800米跑(女)_成绩', 'col_score': '1000米跑(男)/800米跑(女)_分数', 'name': '1000米跑'},
    '800米跑': {'col_grade': '1000米跑(男)/800米跑(女)_成绩', 'col_score': '1000米跑(男)/800米跑(女)_分数', 'name': '800米跑'}
}

TEST_CONFIG = {
    '通用': ['BMI', '肺活量'],
    '小学一二年级': ['50米跑', '坐位体前屈', '1分钟跳绳'],
    '小学三四年级': ['50米跑', '坐位体前屈', '1分钟跳绳', '1分钟仰卧起坐'],
    '小学五六年级': ['50米跑', '坐位体前屈', '1分钟跳绳', '1分钟仰卧起坐', '50米×8往返跑'],
    '中学及以上': {
        '通用': ['50米跑', '坐位体前屈', '立定跳远'],
        '男': ['引体向上', '1000米跑'],
        '女': ['仰卧起坐(女)', '800米跑']
    }
}

# ★ 2. 训练建议库
TRAINING_SUGGESTIONS_DETAILED = {
    "训练总纲": [
        {"title": "安全永远第一", "content": "训练前充分热身（10分钟动态拉伸），训练后整理放松（10分钟静态拉伸）。身体不适时，必须停止训练。"},
        {"title": "循序渐进", "content": "严格按照自己所在阶段的计划进行，不要盲目增加强度，避免受伤。"},
        {"title": "持之以恒", "content": "体育成绩的提升没有捷径，关键在于每周坚持训练3-4次，养成习惯。"},
        {"title": "营养与休息", "content": "保证每天8-9小时的充足睡眠，均衡饮食，多吃蛋白质（牛奶、鸡蛋、肉类）和蔬菜，为身体恢复提供保障。"}
    ],
    "BMI指数": { "通用": { "不及格": "<strong>目标：</strong>调整BMI至及格范围。<br><strong>过重者:</strong> 饮食上减少油炸食品、含糖饮料和零食；运动上每周进行4-5次中等强度的有氧运动（慢跑、游泳），每次30-40分钟。<br><strong>过轻者:</strong> 饮食上适当加餐（牛奶、坚果）；运动上每周进行2-3次力量训练（俯卧撑、深蹲）以增加肌肉重量。", "及格": "<strong>目标：</strong>保持健康的身体状态。<br><strong>计划：</strong>维持均衡饮食和规律运动的良好习惯。", "良好": "<strong>目标：</strong>保持健康的身体状态。<br><strong>计划：</strong>维持均衡饮食和规律运动的良好习惯。", "优秀": "<strong>目标：</strong>保持健康的身体状态。<br><strong>计划：</strong>维持均衡饮食和规律运动的良好习惯。"}},
    "肺活量": { "通用": { "不及格": "<strong>目标：</strong>建立基础心肺功能。<br><strong>计划：</strong>每周3次匀速慢跑（15-20分钟），每天早晚进行5分钟腹式呼吸练习。", "及格": "<strong>目标：</strong>增强心肺耐力。<br><strong>计划：</strong>每周2次变速跑；每周2次跳绳（连续3分钟x6组）。", "良好": "<strong>目标：</strong>挑战心肺极限。<br><strong>计划：</strong>每周1-2次400米间歇跑；每周1次2000-3000米长跑。", "优秀": "<strong>目标：</strong>保持顶尖心肺功能。<br><strong>计划：</strong>每周1次高强度间歇训练（HIIT），或进行游泳、篮球等交叉训练。"}},
    "50米跑": { "男": { "不及格": "<strong>目标：</strong>跑进9.7秒。<br><strong>计划：</strong>每周3次基础力量（深蹲）；每周2次技术模仿（高抬腿）；每周2次启动练习（20米冲刺）。", "及格": "<strong>目标：</strong>跑进8.5秒。<br><strong>计划：</strong>每周2次步频训练（快速小步跑）；每周2次30-50米加速跑。", "良好": "<strong>目标：</strong>跑进7.8秒。<br><strong>计划：</strong>每周2次爆发力训练（蛙跳）；每周2次50-60米全速跑。", "优秀": "<strong>目标：</strong>巩固成绩。<br><strong>计划：</strong>练习听信号起跑；每周1次缓坡冲刺。"}, "女": { "不及格": "<strong>目标：</strong>跑进10.6秒。<br><strong>计划：</strong>每周3次基础力量（深蹲）；每周2次技术模仿（高抬腿）；每周2次启动练习（20米冲刺），重点是克服恐惧，敢于发力。", "及格": "<strong>目标：</strong>跑进9.4秒。<br><strong>计划：</strong>每周2次步频训练（快速小步跑）；每周2次30-50米加速跑，重点是提高摆臂频率和抬腿幅度。", "良好": "<strong>目标：</strong>跑进8.5秒。<br><strong>计划：</strong>每周2次爆发力训练（蛙跳）；每周2次50-60米全速跑，增强力量，提升冲刺速度。", "优秀": "<strong>目标：</strong>保持状态。<br><strong>计划：</strong>练习听信号起跑；每周1次缓坡冲刺，优化技术细节。"}},
    "坐位体前屈": { "通用": { "不及格": "<strong>目标：</strong>达到及格线。<br><strong>计划：</strong>每天进行静态拉伸，尤其在运动后。保持拉伸感30-60秒，重复4-5次。", "及格": "<strong>目标：</strong>达到良好。<br><strong>计划：</strong>加大强度。每次拉伸时间延长至60秒以上，可请同学或用毛巾辅助。", "良好": "<strong>目标：</strong>冲击优秀。<br><strong>计划：</strong>动态与静态结合。训练前进行动态拉伸（如踢腿），训练后进行深度静态拉伸。", "优秀": "<strong>目标：</strong>保持顶级柔韧性。<br><strong>计划：</strong>将拉伸融入日常生活，并可以尝试瑜伽动作，如“下犬式”。"}},
    "立定跳远": { "男": { "不及格": "<strong>目标：</strong>跳过174cm。<br><strong>计划：</strong>每周3次动作分解练习（摆臂下蹲、协调起跳）；每天做3组提踵。", "及格": "<strong>目标：</strong>跳过218cm。<br><strong>计划：</strong>每周2次力量增强（深蹲跳）；每周进行20-30次完整技术练习。", "良好": "<strong>目标：</strong>跳过240cm。<br><strong>计划：</strong>每周2次爆发力强化（蛙跳）；每周3次核心训练（平板支撑）。", "优秀": "<strong>目标：</strong>最大化成绩。<br><strong>计划：</strong>尝试跳台阶或增高跳练习（需指导）；优化落地缓冲技术。"}, "女": { "不及格": "<strong>目标：</strong>跳过147cm。<br><strong>计划：</strong>每周3次动作分解练习（摆臂下蹲、协调起跳）；每天做3组提踵，核心是学会协调发力。", "及格": "<strong>目标：</strong>跳过181cm。<br><strong>计划：</strong>每周2次力量增强（深蹲跳）；每周进行20-30次完整技术练习，提升下肢力量是关键。", "良好": "<strong>目标：</strong>跳过200cm。<br><strong>计划：</strong>每周2次爆发力强化（蛙跳）；每周3次核心训练（平板支撑），加强爆发力和核心。", "优秀": "<strong>目标：</strong>挑战满分。<br><strong>计划：</strong>尝试跳台阶或增高跳练习（需指导），精进技术细节。"}},
    "引体向上": { "男": { "不及格": "<strong>目标：</strong>完成5个以上。<br><strong>计划：</strong>每天练习静力悬挂；练习离心训练（缓慢下降）；使用弹力带辅助。", "及格": "<strong>目标：</strong>完成11个以上。<br><strong>计划：</strong>每周3次分组训练，每组做到力竭的80%；练习反向划船。", "良好": "<strong>目标：</strong>完成14个。<br><strong>计划：</strong>尝试金字塔训练法；提高总训练量。", "优秀": "<strong>目标：</strong>保持高水平。<br><strong>计划：</strong>确保每一个动作都标准；尝试不同握距。"}},
    "仰卧起坐": { "女": { "不及格": "<strong>目标：</strong>达到23个/分钟。<br><strong>计划：</strong>每周3次核心基础（平板支撑、卷腹）；进行不计时的标准动作练习。", "及格": "<strong>目标：</strong>达到44个/分钟。<br><strong>计划：</strong>每周3次耐力训练（4-5组x25个）；每周进行2次30秒快速测试。", "良好": "<strong>目标：</strong>冲击51个/分钟。<br><strong>计划：</strong>进行1分钟模拟测试配速；加入高阶核心动作。", "优秀": "<strong>目标：</strong>确保满分。<br><strong>计划：</strong>进行75-90秒的超量训练；进行全真模拟考试。"}},
    "1000米跑": { "男": { "不及格": "<strong>目标：</strong>跑进5'00\"。<br><strong>计划：</strong>每周2-3次有氧基础跑，不停歇地跑完1500米。", "及格": "<strong>目标：</strong>跑进4'07\"。<br><strong>计划：</strong>每周1次配速感知跑，1000米拆为2个500米。", "良好": "<strong>目标：</strong>跑进3'50\"。<br><strong>计划：</strong>每周1次速度耐力跑，400米间歇跑。", "优秀": "<strong>目标：</strong>突破极限。<br><strong>计划：</strong>在每次耐力跑的最后200米全力冲刺。"}},
    "800米跑": { "女": { "不及格": "<strong>目标：</strong>跑进4'45\"。<br><strong>计划：</strong>每周2-3次慢跑1000-1200米，培养跑步习惯。", "及格": "<strong>目标：</strong>跑进3'52\"。<br><strong>计划：</strong>每周1次配速学习，2组400米跑。", "良好": "<strong>目标：</strong>跑进3'30\"。<br><strong>计划：</strong>每周1次速度提升，200米间歇跑。", "优秀": "<strong>目标：</strong>稳定满分。<br><strong>计划：</strong>每周进行一次800米模拟测试。"}},
    "1分钟跳绳": { "通用": { "不及格": "<strong>目标：</strong>提高协调性和耐力。<br><strong>计划：</strong>从慢速双脚跳开始，熟练后再提速。目标是能连续跳1分钟不停歇。", "及格": "<strong>目标：</strong>减少失误，提高速度。<br><strong>计划：</strong>练习高抬腿跳，并进行30秒快速跳测试。", "良好": "<strong>目标：</strong>冲击更高次数。<br><strong>计划：</strong>尝试更快的速度，并可以学习双摇等更复杂的动作。", "优秀": "<strong>目标：</strong>保持顶尖水平。<br><strong>计划：</strong>将跳绳作为日常热身和心肺训练的一部分。"}},
    "1分钟仰卧起坐": { "通用": { "不及格": "<strong>目标：</strong>掌握正确动作，加强核心。<br><strong>计划：</strong>从基础的卷腹和平板支撑开始，打好核心力量基础。", "及格": "<strong>目标：</strong>提高耐力。<br><strong>计划：</strong>进行多组数练习，例如每组20个，完成4-5组。", "良好": "<strong>目标：</strong>提高速度和数量。<br><strong>计划：</strong>进行30秒或1分钟的模拟测试，找到自己的节奏。", "优秀": "<strong>目标：</strong>保持高水平。<br><strong>计划：</strong>加入两头起、俄罗斯转体等更高阶的核心训练动作。"}},
    "50米×8往返跑": { "通用": { "不及格": "<strong>目标：</strong>提高灵敏性和转身技巧。<br><strong>计划：</strong>多进行短距离的折返跑练习，熟悉转身和身体的启动、制动。", "及格": "<strong>目标：</strong>提升无氧耐力。<br><strong>计划：</strong>进行完整的模拟测试，注意呼吸节奏的调整。", "良好": "<strong>目标：</strong>缩短总用时。<br><strong>计划：</strong>在训练中加入变向跑、交叉步等灵敏性训练。", "优秀": "<strong>目标：</strong>保持高水平。<br><strong>计划：</strong>将往返跑作为一项综合性的速度和耐力训练。"}}
}

# ★ 3. 权重规则库
WEIGHT_CONFIG = {
    '通用': {'BMI': 0.15, '肺活量': 0.15},
    '小学一二年级': {'50米跑': 0.20, '坐位体前屈': 0.30, '1分钟跳绳': 0.20},
    '小学三四年级': {'50米跑': 0.20, '坐位体前屈': 0.20, '1分钟跳绳': 0.20, '1分钟仰卧起坐': 0.10},
    '小学五六年级': {'50米跑': 0.20, '坐位体前屈': 0.10, '1分钟跳绳': 0.10, '1分钟仰卧起坐': 0.20, '50米×8往返跑': 0.10},
    '中学及以上': {'50米跑': 0.20, '坐位体前屈': 0.10, '立定跳远': 0.10, '引体向上': 0.10, '仰卧起坐(女)': 0.10, '1000米跑': 0.20, '800米跑': 0.20}
}

# --- S2 常量 ---
TOTAL_SCORE_CATEGORIES_S2 = {
    '90分优秀': {'excellent_threshold': 90, 'good_threshold_min': 80, 'good_threshold_max': 89.999, 'pass_threshold_min': 60, 'pass_threshold_max': 79.999, 'excellent_good_rate_threshold': 80},
    '72分优秀': {'excellent_threshold': 72, 'good_threshold_min': 64, 'good_threshold_max': 71.999, 'pass_threshold_min': 48, 'pass_threshold_max': 63.999, 'excellent_good_rate_threshold': 64}
}
REGION_MAP_S2 = {'371523': '茌平区', '371524': '东阿县', '371502': '东昌府区', '371593': '度假区', '371526': '高唐县', '371592': '高新区', '371525': '冠县', '371591': '开发区', '371581': '临清市', '371522': '莘县', '371521': '阳谷县'}
log_messages = []

# ================================================================================
# ================================================================================
#  SCRIPT 1: 个人报告生成器 (核心后端函数)
# ================================================================================
# ================================================================================

def process_complex_header(df):
    """处理个人报告生成器上传的复杂合并单元格表头"""
    if df.shape[0] < 2:
        st.error("错误：文件内容不足两行，无法解析两级表头。")
        return None
    try:
        header_row1 = df.iloc[0].ffill()
        header_row2 = df.iloc[1].fillna('')
        new_columns = [f"{c1}_{c2}".strip('_') if c1 and not str(c1).startswith('Unnamed') else str(c2) for c1, c2 in zip(header_row1, header_row2)]
        df.columns = new_columns
        df = df.iloc[2:].reset_index(drop=True)
        return df
    except Exception as e:
        st.error(f"处理表头时发生严重错误: {e}")
        return None

def extract_grade_from_class(class_name):
    """从班级名称中智能提取年级信息 (增强版)"""
    if not isinstance(class_name, str):
        return "未知"
    if '初' in class_name: return "初中年级"
    if '高' in class_name: return "高中年级"
    if '大' in class_name: return "大学年级"
    match = re.search(r'([一二三四五六])年级', class_name)
    if match:
        return f"{match.group(1)}年级"
    return "未知"

def get_applicable_tests(grade, gender):
    """根据学生的年级和性别，返回其应测的项目字典 (修正版)"""
    applicable_keys = set(TEST_CONFIG['通用'])
    grade_map = {'一年级': '小学一二年级', '二年级': '小学一二年级', '三年级': '小学三四年级', '四年级': '小学三四年级', '五年级': '小学五六年级', '六年级': '小学五六年级'}
    matched_key = next((grade_map[key] for key in grade_map if key in grade), None)
    
    if matched_key:
        applicable_keys.update(TEST_CONFIG[matched_key])
    elif any(keyword in grade for keyword in ['初', '高', '大']):
        config = TEST_CONFIG['中学及以上']
        applicable_keys.update(config['通用'])
        applicable_keys.update(config['男'] if gender == '男' else config['女'])
    
    return [ALL_TESTS[key] for key in ALL_TESTS if key in applicable_keys]

def get_rating(score):
    """(保持不变) 根据100分制分数获取单项评级"""
    score = pd.to_numeric(score, errors='coerce')
    if pd.isna(score): return "未知"
    RATING_STANDARDS = {(90, 101): "优秀", (80, 90): "良好", (60, 80): "及格", (0, 60): "不及格"}
    return next((rating for score_range, rating in RATING_STANDARDS.items() if score_range[0] <= score < score_range[1]), "未知")

def get_rating_30(score):
    """【新增】根据30分制分数获取综合评级"""
    score = pd.to_numeric(score, errors='coerce')
    if pd.isna(score): return "未知"
    RATING_STANDARDS_30 = {(27, 31): "优秀", (24, 27): "良好", (18, 24): "及格", (0, 18): "不及格"}
    return next((rating for score_range, rating in RATING_STANDARDS_30.items() if score_range[0] <= score < score_range[1]), "未知")

def calculate_and_apply_weights(df):
    """
    【最终生产版 - 正确应用30分制】
    1. 计算100分制下的加权单项分。
    2. 将100分制的加权单项分求和得到100分制的加权总分。
    3. 将所有分数（单项和总分）按比例缩放至30分制。
    """
    df_new = df.copy()
    
    score_cols_in_df = [col for col in df.columns if isinstance(col, str) and '_分数' in col]

    # 步骤1 & 2: 计算100分制下的加权分数和总分
    for index, student_data in df.iterrows():
        grade = student_data.get('年级', '未知')
        gender = student_data.get('性别', '')
        
        student_weight_config = WEIGHT_CONFIG['通用'].copy()
        grade_map = { '一年级': '小学一二年级', '二年级': '小学一二年级', '三年级': '小学三四年级', '四年级': '小学三四年级', '五年级': '小学五六年级', '六年级': '小学五六年级'}
        matched_grade_key = next((grade_map[key] for key in grade_map if key in grade), None)
        if matched_grade_key:
            student_weight_config.update(WEIGHT_CONFIG[matched_grade_key])
        elif any(keyword in grade for keyword in ['初', '高', '大']):
            student_weight_config.update(WEIGHT_CONFIG['中学及以上'])

        total_weighted_score_100 = 0.0
        
        for score_col in score_cols_in_df:
            current_test_key = next((key for key, value in ALL_TESTS.items() if value['col_score'] == score_col), None)
            
            weight = 0.0
            if current_test_key and current_test_key in student_weight_config:
                is_gender_match = not (
                    (current_test_key == '仰卧起坐(女)' and gender == '男') or
                    (current_test_key == '引体向上' and gender == '女') or
                    (current_test_key == '1000米跑' and gender == '女') or
                    (current_test_key == '800米跑' and gender == '男')
                )
                if is_gender_match:
                    weight = student_weight_config[current_test_key]

            original_score = pd.to_numeric(student_data.get(score_col), errors='coerce')
            original_score = original_score if pd.notna(original_score) else 0.0
            
            weighted_score_100 = original_score * weight
            total_weighted_score_100 += weighted_score_100
            
            df_new.loc[index, score_col] = weighted_score_100

        df_new.loc[index, '总分'] = total_weighted_score_100

    # 步骤3: 将所有分数（单项和总分）按比例缩放至30分制
    for col in score_cols_in_df + ['总分']:
        if col in df_new.columns:
            df_new[col] = df_new[col] * 0.3
            
    return df_new

def create_radar_chart(student_data, test_items, student_weight_config):
    """【最终雷达图修正版 - 归一化显示】"""
    score_items = [item for item in test_items if 'col_score' in item and item['col_score'] in student_data]
    if not score_items: return None

    labels = []
    normalized_stats = []

    for item in score_items:
        current_test_key = next((key for key, value in ALL_TESTS.items() if value['name'] == item['name']), None)
        
        if current_test_key and current_test_key in student_weight_config:
            weight = student_weight_config[current_test_key]
            max_weighted_score = 100 * weight * 0.3
            actual_weighted_score = pd.to_numeric(student_data.get(item['col_score'], 0), errors='coerce')
            actual_weighted_score = actual_weighted_score if pd.notna(actual_weighted_score) else 0
            if max_weighted_score > 0:
                normalized_score = (actual_weighted_score / max_weighted_score) * 100
                normalized_stats.append(min(normalized_score, 100))
                labels.append(item['name'])
    
    if not labels or not normalized_stats: return None

    labels_closed = labels + [labels[0]]
    stats_closed = normalized_stats + [normalized_stats[0]]
    
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(r=[100] * len(labels_closed), theta=labels_closed, fill='toself', fillcolor='rgba(0, 150, 0, 0.1)', line=dict(color='rgba(0,0,0,0)'), hoverinfo='none', name='优秀区'))
    fig.add_trace(go.Scatterpolar(r=[90] * len(labels_closed), theta=labels_closed, fill='toself', fillcolor='rgba(0, 100, 200, 0.2)', line=dict(color='rgba(0,0,0,0)'), hoverinfo='none', name='良好区'))
    fig.add_trace(go.Scatterpolar(r=[80] * len(labels_closed), theta=labels_closed, fill='toself', fillcolor='rgba(255, 140, 0, 0.3)', line=dict(color='rgba(0,0,0,0)'), hoverinfo='none', name='及格区'))
    fig.add_trace(go.Scatterpolar(r=[60] * len(labels_closed), theta=labels_closed, fill='toself', fillcolor='rgba(220, 20, 60, 0.2)', line=dict(color='rgba(0,0,0,0)'), hoverinfo='none', name='不及格区'))
    fig.add_trace(go.Scatterpolar(r=stats_closed, theta=labels_closed, fill='toself', fillcolor='rgba(22, 118, 248, 0.4)', line=dict(color='#1f77b4', width=3), name='学生表现', hovertemplate='<b>%{theta}</b><br>表现水平: %{r:.1f}/100<extra></extra>'))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100], showticklabels=False, ticks='', gridcolor='rgba(0,0,0,0.1)'), angularaxis=dict(gridcolor='rgba(0,0,0,0.1)')), 
        showlegend=False, font=dict(family="sans-serif", size=16, color="#333"), margin=dict(l=80, r=80, t=40, b=40), 
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
    return fig

def generate_pdf_report(student_data, radar_image_bytes, test_items):
    """
    【最终重构版 - 适用于 fpdf2】
    """
    THEME_COLOR = (22, 118, 248); LIGHT_THEME_COLOR = (230, 240, 255); TEXT_COLOR = (50, 50, 50); LIGHT_TEXT_COLOR = (120, 120, 120)
    RATING_COLORS = {"优秀": (34, 139, 34), "良好": (0, 100, 200), "及格": (255, 140, 0), "不及格": (220, 20, 60), "未知": (100, 100, 100), "——": (100, 100, 100)}

    class PDF(FPDF):
        def __init__(self, student_name, class_name, test_date):
            super().__init__()
            self.student_name = str(student_name)
            self.class_name   = str(class_name)
            self.test_date    = str(test_date)

            # 字体路径
            font_path = ASSETS_DIR / "fonts" / "SimHei.ttf"
            if font_path.exists():
                self.add_font('SimHei', '', str(font_path))
                self.font_family = 'SimHei'
            else:
                self.font_family = 'Helvetica'
                if 'font_warning_sent' not in st.session_state:
                    st.warning("【重要】未找到 SimHei.ttf 字体文件，PDF中的中文将是乱码。")
                    st.session_state.font_warning_sent = True

            # 必须放在方法里！
            self.set_auto_page_break(auto=True, margin=15)

        def header(self):
            if self.page_no() == 1: return
            self.set_font(self.font_family, '', 9)
            self.set_text_color(120, 120, 120)
            self.cell(0, 10, f"{self.student_name} | {self.class_name}", 0, 0, 'L')
            self.cell(0, 10, f"报告生成日期: {self.test_date}", 0, 1, 'R')
            self.set_draw_color(220, 220, 220); self.line(10, self.get_y(), 200, self.get_y()); self.ln(5)

        def footer(self):
            if self.page_no() > 1:
                self.set_y(-15)
                self.set_draw_color(22, 118, 248); self.set_line_width(0.5)
                self.line(10, self.get_y() - 2, 200, self.get_y() - 2)
                self.set_font(self.font_family, '', 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 10, f'第 {self.page_no() - 1} 页', 0, 0, 'C')

        def section_card(self, title, icon, content_func):
            if self.get_y() > 250: self.add_page()
            self.set_font(self.font_family, '', 16)
            self.set_text_color(22, 118, 248)
            self.cell(0, 12, f"{icon} {title}", 0, 1, 'L')
            self.set_draw_color(240, 240, 240); self.line(10, self.get_y(), 200, self.get_y()); self.ln(5)
            self.set_text_color(50, 50, 50)
            content_func(self)
            self.ln(10)
    
    try:
        pdf = PDF(student_data.get('姓名', 'N/A'), student_data.get('班级', 'N/A'), pd.Timestamp.now().strftime("%Y-%m-%d"))
        pdf.add_page(); pdf.set_fill_color(230, 240, 255); pdf.rect(0, 0, 210, 297, 'F')
        
        logo_path = ASSETS_DIR / "images" / "logo.png"
        if logo_path.exists():
            try:
                pdf.image(str(logo_path), x=85, y=40, w=40)
            except Exception:
                pass
            
        pdf.set_y(90); pdf.set_font(pdf.font_family, '', 36); pdf.set_text_color(22, 118, 248); pdf.cell(0, 20, '个人体质健康报告', 0, 1, 'C')
        pdf.set_y(150); pdf.set_font(pdf.font_family, '', 18); pdf.set_text_color(50, 50, 50)
        pdf.cell(0, 12, f"姓名: {student_data.get('姓名', 'N/A')}", 0, 1, 'C')
        pdf.cell(0, 12, f"班级: {student_data.get('班级', 'N/A')}", 0, 1, 'C')
        pdf.cell(0, 12, f"学籍号: {student_data.get('学籍号', 'N/A')}", 0, 1, 'C')
        pdf.set_y(250); pdf.set_font(pdf.font_family, '', 10); pdf.set_text_color(120, 120, 120); pdf.cell(0, 10, "报告生成于 " + pd.Timestamp.now().strftime("%Y年%m月%d日"), 0, 1, 'C')
        pdf.add_page()
        
        def content_summary(p):
            initial_y = p.get_y()
            total_score_val = pd.to_numeric(student_data.get('总分'), errors='coerce')
            total_score = total_score_val if pd.notna(total_score_val) else 0
            total_rating = get_rating_30(total_score)
            p.set_font(p.font_family, '', 12)
            p.multi_cell(w=80, h=8, txt=f"本次体测总分为: {total_score:.2f} 分 (30分制)，综合评级为: {total_rating}。", border=0, align='L')
            if radar_image_bytes and hasattr(radar_image_bytes, 'read'):
                try:
                    radar_image_bytes.name = 'radar.png'; p.image(radar_image_bytes, x=p.get_x() + 0, y=initial_y - 5, w=80)
                except Exception as e:
                    p.set_font(p.font_family, '', 8); p.set_text_color(255, 0, 0)
                    p.text(p.get_x() + 80, initial_y, f"雷达图生成失败"); p.set_text_color(0, 0, 0)
            p.set_y(initial_y + 85)
        
        def content_details(p):
            p.set_font(p.font_family, '', 11); p.set_fill_color(240, 240, 240); p.set_text_color(22, 118, 248)
            p.cell(50, 10, '测试项目', 1, 0, 'C', fill=True); p.cell(40, 10, '你的成绩', 1, 0, 'C', fill=True); p.cell(40, 10, '加权得分', 1, 0, 'C', fill=True); p.cell(40, 10, '单项评级', 1, 1, 'C', fill=True)
            p.set_text_color(50, 50, 50)
            for item in test_items:
                grade = student_data.get(item.get('col_grade'), 'N/A'); score = student_data.get(item.get('col_score'), 'N/A')
                rating_col_name = item.get('col_score', '').replace('_分数', '_评级')
                rating = student_data.get(rating_col_name, "——")
                p.cell(50, 10, str(item['name']), 1, 0, 'C')
                p.cell(40, 10, str(grade), 1, 0, 'C')
                p.cell(40, 10, str(f"{pd.to_numeric(score, errors='coerce'):.2f}" if pd.notna(pd.to_numeric(score, errors='coerce')) else score), 1, 0, 'C')
                p.set_text_color(*RATING_COLORS.get(rating, (0,0,0))); p.cell(40, 10, str(rating), 1, 1, 'C'); p.set_text_color(50, 50, 50)

        def content_suggestions(p):
            p.set_font(p.font_family, '', 11); t_g_text = "\n".join([f"• {item['title']}: {item['content']}" for item in TRAINING_SUGGESTIONS_DETAILED['训练总纲']]); p.multi_cell(0, 6, "训练总纲 (必读):\n" + t_g_text, "T", 'L'); p.ln(5)
            gender = student_data.get('性别', '男')
            suggestions = {}
            for item in test_items:
                if 'col_score' in item:
                    rating = student_data.get(item['col_score'].replace('_分数', '_评级'), "未知")
                    if rating not in ["未知", "——"]:
                        text = TRAINING_SUGGESTIONS_DETAILED.get(item['name'], {}).get(gender, TRAINING_SUGGESTIONS_DETAILED.get(item['name'], {}).get("通用", {})).get(rating, "")
                        if text: suggestions[item['name']] = {"rating": rating, "text": text}
            if not suggestions:
                p.set_font(p.font_family, '', 11); p.set_text_color(34, 139, 34); p.multi_cell(0, 8, "太棒了！所有项目均表现良好或优秀，请继续保持！", 0, 'L')
            else:
                p.set_left_margin(10)
                for cat, info in suggestions.items():
                    p.set_font(p.font_family, '', 12); p.set_text_color(22, 118, 248); p.cell(0, 8, f"• 针对「{cat}」(评级: {info['rating']}) 的反馈与建议:", 0, 1, 'L')
                    p.set_text_color(50, 50, 50); p.set_font(p.font_family, '', 11); p.set_left_margin(15); p.multi_cell(0, 7, info['text'].replace('<br>', '\n').replace('<strong>', '').replace('</strong>', ''), 0, 'L'); p.set_left_margin(10); p.ln(3)
        
        pdf.section_card("📊 核心指标概览", "📊", content_summary)
        pdf.section_card("📈 各项指标详解", "📈", content_details)
        pdf.add_page()
        pdf.section_card("🏃 个性化训练计划", "🏃", content_suggestions)
        appendix_image_path = str(ASSETS_DIR / "images" / "pkffbnvy.png")
        if os.path.exists(appendix_image_path):
            try:
                pdf.add_page()
                pdf.section_card("📎 附录：评分标准参考", "📎", lambda p: p.image(appendix_image_path, w=180))
            except Exception:
                pass
        
        return bytes(pdf.output())

    except Exception as e:
        st.error(f"生成PDF时发生严重错误: {e}")
        st.exception(e)
        return None

# ================================================================================
# ================================================================================
# SCRIPT 2: 数据统计与工具箱 (核心后端函数)
# ================================================================================
# ================================================================================

def s2_log(message, clear=False):
    """记录日志消息以供显示"""
    if clear:
        log_messages.clear()
    log_messages.append(message)

def sanitize_sheet_name(name):
    """清洁字符串为有效的Excel表名。"""
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars: name = name.replace(char, '-')
    return name[:31]

def sanitize_filename(name):
    """清洁字符串为有效的文件名。"""
    if name is None: return ''
    return "".join(c for c in str(name) if c.isalnum() or c in (' ', '_', '-')).rstrip()

# --- S2 数据统计分析核心函数 ---

def _s2_add_region_column(df):
    if '学籍号' not in df.columns:
        s2_log("警告: 未找到'学籍号'列，无法添加区域信息。")
        df['区域'] = '未找到学籍号'
        return df
    df['区域'] = df['学籍号'].astype(str).str[4:10].map(REGION_MAP_S2).fillna('未知区域')
    return df

def _s2_extract_grade_and_stage(df):
    if '班级' not in df.columns:
        s2_log("警告: 缺少“班级”列，无法提取学段和年级信息。")
        df['学段'] = '未知'; df['年级'] = 0
        return df
    def get_info(class_name):
        class_name = str(class_name)
        if match := re.search(r'高([一二三])', class_name): return '高中', {'一': 10, '二': 11, '三': 12}.get(match.group(1), 0)
        if match := re.search(r'初([一二三四])', class_name): return '初中', {'一': 7, '二': 8, '三': 9, '四': 10}.get(match.group(1), 0)
        if match := re.search(r'([一二三四五六])年级', class_name): return '小学', {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6}.get(match.group(1), 0)
        return '未知', 0
    df[['学段', '年级']] = df['班级'].apply(lambda x: pd.Series(get_info(x)))
    s2_log("已从“班级”列中提取学段和年级信息。")
    return df

def _s2_calculate_group_total_score_stats(df, group_by_cols, score_thresholds, include_total_row=False):
    """
    按指定列对总分进行分组统计。
    """
    if not isinstance(group_by_cols, list): 
        group_by_cols = [group_by_cols]
    
    primary_group_cols = [col for col in group_by_cols if col != '性别']
    
    all_group_stats = []
    
    # 如果有主分组（非性别），则按主分组进行外层循环
    outer_grouped = df.groupby(primary_group_cols, dropna=False) if primary_group_cols else [((), df)]

    for outer_group_keys, outer_group_df in outer_grouped:
        # 在每个主分组内，再按性别进行内层循环
        inner_grouped_by_gender = outer_group_df.groupby('性别', as_index=False, dropna=False)
        
        for gender_key, gender_group in inner_grouped_by_gender:
            total_students = len(gender_group)
            if total_students == 0: 
                continue
            
            # 确保总分是数值类型，无效值填充为0
            tested_scores_group = pd.to_numeric(gender_group['总分'], errors='coerce').fillna(0)
            
            excellent_count = (tested_scores_group >= score_thresholds['excellent_threshold']).sum()
            good_count = ((tested_scores_group >= score_thresholds['good_threshold_min']) & (tested_scores_group <= score_thresholds['good_threshold_max'])).sum()
            pass_count = ((tested_scores_group >= score_thresholds['pass_threshold_min']) & (tested_scores_group <= score_thresholds['pass_threshold_max'])).sum()
            fail_count = (tested_scores_group < score_thresholds['pass_threshold_min']).sum()

            row_data = {
                '总人数': total_students, 
                '实际测试人数': total_students,
                f"总分满分人数 (100分)": (tested_scores_group == 100).sum(),
                f"总分达到{score_thresholds['excellent_threshold']}分优秀人数": excellent_count,
                f"总分良好人数 ({int(score_thresholds['good_threshold_min'])}-{int(score_thresholds['good_threshold_max'])}分)": good_count,
                f"总分及格人数 ({int(score_thresholds['pass_threshold_min'])}-{int(score_thresholds['pass_threshold_max'])}分)": pass_count,
                f"总分{int(score_thresholds['pass_threshold_min'])}分以下不及格人数": fail_count,
                '总分合格人数': (tested_scores_group >= score_thresholds['pass_threshold_min']).sum(),
                '总分平均分': round(tested_scores_group.mean(), 2) if total_students > 0 else 0,
                '总分最高分': round(tested_scores_group.max(), 2) if total_students > 0 else 0,
                '总分最低分': round(tested_scores_group.min(), 2) if total_students > 0 else 0,
                f"总分达到{score_thresholds['excellent_threshold']}分优秀率": round(excellent_count / total_students, 4) if total_students > 0 else 0,
                f"总分良好率 ({int(score_thresholds['good_threshold_min'])}-{int(score_thresholds['good_threshold_max'])}分)": round(good_count / total_students, 4) if total_students > 0 else 0,
                f"总分及格率 ({int(score_thresholds['pass_threshold_min'])}-{int(score_thresholds['pass_threshold_max'])}分)": round(pass_count / total_students, 4) if total_students > 0 else 0,
                f"总分{int(score_thresholds['pass_threshold_min'])}分以下不及格率": round(fail_count / total_students, 4) if total_students > 0 else 0,
                '总分合格率': round((tested_scores_group >= score_thresholds['pass_threshold_min']).sum() / total_students, 4) if total_students > 0 else 0,
                '优良率': round((tested_scores_group > score_thresholds['excellent_good_rate_threshold']).sum() / total_students, 4) if total_students > 0 else 0,
                '测试率': 1.0
            }
            
            full_row = {}
            if primary_group_cols:
                keys = outer_group_keys if isinstance(outer_group_keys, tuple) else (outer_group_keys,)
                full_row.update(dict(zip(primary_group_cols, keys)))
            full_row['性别'] = gender_key
            full_row.update(row_data)
            all_group_stats.append(full_row)
        
        # 如果需要，为每个主分组计算总计行
        if include_total_row and not outer_group_df.empty:
            total_students = len(outer_group_df)
            tested_scores_group = pd.to_numeric(outer_group_df['总分'], errors='coerce').fillna(0)
            excellent_count = (tested_scores_group >= score_thresholds['excellent_threshold']).sum()
            good_count = ((tested_scores_group >= score_thresholds['good_threshold_min']) & (tested_scores_group <= score_thresholds['good_threshold_max'])).sum()
            pass_count = ((tested_scores_group >= score_thresholds['pass_threshold_min']) & (tested_scores_group <= score_thresholds['pass_threshold_max'])).sum()
            fail_count = (tested_scores_group < score_thresholds['pass_threshold_min']).sum()

            total_row_data = {
                '总人数': total_students, 
                '实际测试人数': total_students,
                f"总分满分人数 (100分)": (tested_scores_group == 100).sum(),
                f"总分达到{score_thresholds['excellent_threshold']}分优秀人数": excellent_count,
                f"总分良好人数 ({int(score_thresholds['good_threshold_min'])}-{int(score_thresholds['good_threshold_max'])}分)": good_count,
                f"总分及格人数 ({int(score_thresholds['pass_threshold_min'])}-{int(score_thresholds['pass_threshold_max'])}分)": pass_count,
                f"总分{int(score_thresholds['pass_threshold_min'])}分以下不及格人数": fail_count,
                '总分合格人数': (tested_scores_group >= score_thresholds['pass_threshold_min']).sum(),
                '总分平均分': round(tested_scores_group.mean(), 2) if total_students > 0 else 0,
                '总分最高分': round(tested_scores_group.max(), 2) if total_students > 0 else 0,
                '总分最低分': round(tested_scores_group.min(), 2) if total_students > 0 else 0,
                f"总分达到{score_thresholds['excellent_threshold']}分优秀率": round(excellent_count / total_students, 4) if total_students > 0 else 0,
                f"总分良好率 ({int(score_thresholds['good_threshold_min'])}-{int(score_thresholds['good_threshold_max'])}分)": round(good_count / total_students, 4) if total_students > 0 else 0,
                f"总分及格率 ({int(score_thresholds['pass_threshold_min'])}-{int(score_thresholds['pass_threshold_max'])}分)": round(pass_count / total_students, 4) if total_students > 0 else 0,
                f"总分{int(score_thresholds['pass_threshold_min'])}分以下不及格率": round(fail_count / total_students, 4) if total_students > 0 else 0,
                '总分合格率': round((tested_scores_group >= score_thresholds['pass_threshold_min']).sum() / total_students, 4) if total_students > 0 else 0,
                '优良率': round((tested_scores_group > score_thresholds['excellent_good_rate_threshold']).sum() / total_students, 4) if total_students > 0 else 0,
                '测试率': 1.0
            }
            
            full_total_row = {}
            if primary_group_cols:
                keys = outer_group_keys if isinstance(outer_group_keys, tuple) else (outer_group_keys,)
                full_total_row.update(dict(zip(primary_group_cols, keys)))
            full_total_row['性别'] = '总计'
            full_total_row.update(total_row_data)
            all_group_stats.append(full_total_row)
            
    return pd.DataFrame(all_group_stats)

def _s2_calculate_group_indicator_stats(df, group_by_cols, indicator_name, score_col_name, full_score, include_total_row=False):
    """
    按指定列对单个项目分数进行分组统计。
    """
    if not isinstance(group_by_cols, list):
        group_by_cols = [group_by_cols]

    primary_group_cols = [col for col in group_by_cols if col != '性别']
    all_group_stats = []
    
    outer_grouped = df.groupby(primary_group_cols, dropna=False) if primary_group_cols else [((), df)]

    for outer_group_keys, outer_group_df in outer_grouped:
        inner_grouped_by_gender = outer_group_df.groupby('性别', as_index=False, dropna=False)
        
        for gender_key, gender_group in inner_grouped_by_gender:
            total_students = len(gender_group)
            if total_students == 0:
                continue
            
            # 确保分数列是数值类型
            tested_scores_group = pd.to_numeric(gender_group[score_col_name], errors='coerce').fillna(0)
            
            excellent_count = (tested_scores_group >= 90).sum()
            good_count = ((tested_scores_group >= 80) & (tested_scores_group <= 89.999)).sum()
            pass_count = ((tested_scores_group >= 60) & (tested_scores_group <= 79.999)).sum()
            fail_count = (tested_scores_group < 60).sum()
            
            row_data = {
                '总人数': total_students,
                '实际测试人数': total_students,
                f'{indicator_name}满分人数 ({full_score}分)': (tested_scores_group == full_score).sum(),
                f'{indicator_name}达到90分优秀人数': excellent_count,
                f'{indicator_name}良好人数 (80-89分)': good_count,
                f'{indicator_name}及格人数 (60-79分)': pass_count,
                f'{indicator_name}60分以下不及格人数': fail_count,
                f'{indicator_name}合格人数': (tested_scores_group >= 60).sum(),
                f'{indicator_name}平均分': round(tested_scores_group.mean(), 2) if total_students > 0 else 0,
                f'{indicator_name}最高分': round(tested_scores_group.max(), 2) if total_students > 0 else 0,
                f'{indicator_name}最低分': round(tested_scores_group.min(), 2) if total_students > 0 else 0,
                f'{indicator_name}达到90分优秀率': round(excellent_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}良好率 (80-89分)': round(good_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}及格率 (60-79分)': round(pass_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}60分以下不及格率': round(fail_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}合格率': round((tested_scores_group >= 60).sum() / total_students, 4) if total_students > 0 else 0,
                '测试率': 1.0
            }
            
            full_row = {}
            if primary_group_cols:
                keys = outer_group_keys if isinstance(outer_group_keys, tuple) else (outer_group_keys,)
                full_row.update(dict(zip(primary_group_cols, keys)))
            full_row['性别'] = gender_key
            full_row.update(row_data)
            all_group_stats.append(full_row)
            
        if include_total_row and not outer_group_df.empty:
            total_students = len(outer_group_df)
            tested_scores_group = pd.to_numeric(outer_group_df[score_col_name], errors='coerce').fillna(0)
            excellent_count = (tested_scores_group >= 90).sum()
            good_count = ((tested_scores_group >= 80) & (tested_scores_group <= 89.999)).sum()
            pass_count = ((tested_scores_group >= 60) & (tested_scores_group <= 79.999)).sum()
            fail_count = (tested_scores_group < 60).sum()

            total_row_data = {
                '总人数': total_students,
                '实际测试人数': total_students,
                f'{indicator_name}满分人数 ({full_score}分)': (tested_scores_group == full_score).sum(),
                f'{indicator_name}达到90分优秀人数': excellent_count,
                f'{indicator_name}良好人数 (80-89分)': good_count,
                f'{indicator_name}及格人数 (60-79分)': pass_count,
                f'{indicator_name}60分以下不及格人数': fail_count,
                f'{indicator_name}合格人数': (tested_scores_group >= 60).sum(),
                f'{indicator_name}平均分': round(tested_scores_group.mean(), 2) if total_students > 0 else 0,
                f'{indicator_name}最高分': round(tested_scores_group.max(), 2) if total_students > 0 else 0,
                f'{indicator_name}最低分': round(tested_scores_group.min(), 2) if total_students > 0 else 0,
                f'{indicator_name}达到90分优秀率': round(excellent_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}良好率 (80-89分)': round(good_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}及格率 (60-79分)': round(pass_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}60分以下不及格率': round(fail_count / total_students, 4) if total_students > 0 else 0,
                f'{indicator_name}合格率': round((tested_scores_group >= 60).sum() / total_students, 4) if total_students > 0 else 0,
                '测试率': 1.0
            }

            full_total_row = {}
            if primary_group_cols:
                keys = outer_group_keys if isinstance(outer_group_keys, tuple) else (outer_group_keys,)
                full_total_row.update(dict(zip(primary_group_cols, keys)))
            full_total_row['性别'] = '总计'
            full_total_row.update(total_row_data)
            all_group_stats.append(full_total_row)
            
    return pd.DataFrame(all_group_stats)

def _s2_format_excel_file(excel_buffer, total_score_thresholds, mapped_indicator_flat_cols):
    """
    对内存中的Excel文件进行格式化美化。
    """
    s2_log("开始格式化Excel内存文件...")
    try:
        # 从内存中加载工作簿
        workbook = load_workbook(excel_buffer)
        
        # 定义通用样式
        header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        data_font = Font(name='微软雅黑', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # 设置列宽和行高
            for i, col in enumerate(ws.columns):
                ws.column_dimensions[get_column_letter(i + 1)].width = 18 # 统一设置一个较宽的宽度
            ws.row_dimensions[1].height = 40

            # 格式化表头
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 格式化数据行
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    # 根据表头内容设置数字格式
                    header_value = str(ws.cell(row=1, column=cell.column).value)
                    if '率' in header_value:
                        cell.number_format = numbers.FORMAT_PERCENTAGE_00
                    elif '平均分' in header_value or '最高分' in header_value or '最低分' in header_value:
                        cell.number_format = numbers.FORMAT_NUMBER_00
                    elif '人数' in header_value or '总人数' in header_value:
                        cell.number_format = numbers.FORMAT_NUMBER
            
            # 冻结首行
            if ws.max_row > 1:
                ws.freeze_panes = 'A2'

            # 添加条件格式
            if '总分' in sheet_name:
                excellent_threshold = total_score_thresholds['excellent_threshold']
                avg_score_col, excellent_rate_col = None, None
                for col_idx, col_header in enumerate(ws[1], 1):
                    if col_header.value and '平均分' in col_header.value:
                        avg_score_col = col_idx
                    if col_header.value == f'总分达到{excellent_threshold}分优秀率':
                        excellent_rate_col = col_idx
                
                if avg_score_col and ws.max_row > 1:
                    ws.conditional_formatting.add(f'{get_column_letter(avg_score_col)}2:{get_column_letter(avg_score_col)}{ws.max_row}', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color="4BACC6", showValue=True))
                if excellent_rate_col and ws.max_row > 1:
                    ws.conditional_formatting.add(f'{get_column_letter(excellent_rate_col)}2:{get_column_letter(excellent_rate_col)}{ws.max_row}', ColorScaleRule(start_type='num', start_value=0, start_color='FFFFC7CE', mid_type='num', mid_value=0.7, mid_color='FFFFEBB5', end_type='num', end_value=1, end_color='FFC6EFCE'))
            
            for indicator_key in mapped_indicator_flat_cols.keys():
                indicator_name = ALL_TESTS[indicator_key]['name']
                if indicator_name in sheet_name:
                    excellent_rate_col = None
                    for col_idx, col_header in enumerate(ws[1], 1):
                        if col_header.value == f"{indicator_name}达到90分优秀率":
                            excellent_rate_col = col_idx
                            break
                    if excellent_rate_col and ws.max_row > 1:
                         ws.conditional_formatting.add(f'{get_column_letter(excellent_rate_col)}2:{get_column_letter(excellent_rate_col)}{ws.max_row}', ColorScaleRule(start_type='num', start_value=0, start_color='FFFFC7CE', mid_type='num', mid_value=0.7, mid_color='FFFFEBB5', end_type='num', end_value=1, end_color='FFC6EFCE'))
        
        s2_log("Excel格式化成功。")
        # 将格式化后的工作簿保存到一个新的内存流中
        final_buffer = io.BytesIO()
        workbook.save(final_buffer)
        final_buffer.seek(0)
        return final_buffer
    except Exception as e:
        s2_log(f"格式化Excel文件时发生错误: {e}")
        s2_log(traceback.format_exc())
        # 如果格式化失败，返回原始的、未格式化的内存流
        excel_buffer.seek(0)
        return excel_buffer

def _s2_perform_analysis_and_save(df_to_analyze, group_by_selection, score_thresholds, mapped_indicator_flat_cols):
    """
    执行核心分析并将所有结果表格写入一个内存中的Excel文件。
    """
    s2_log("开始生成所有统计表格...")
    base_group_cols = []
    if group_by_selection == '学校和班级': 
        base_group_cols = ['学校', '班级']
    elif group_by_selection == '学校': 
        base_group_cols = ['学校']
    elif group_by_selection == '班级': 
        base_group_cols = ['班级']

    # 使用 BytesIO 作为虚拟的Excel文件在内存中操作
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        
        # --- 总体统计 ---
        s2_log("创建总体总分_性别统计表...")
        overall_total_score_gender_stats = _s2_calculate_group_total_score_stats(df_to_analyze, ['性别'], score_thresholds, include_total_row=True)
        if not overall_total_score_gender_stats.empty:
            overall_total_score_gender_stats.to_excel(writer, sheet_name='总体总分_性别统计', index=False)

        for indicator_key, score_col_name in mapped_indicator_flat_cols.items():
            indicator_name = ALL_TESTS[indicator_key]['name']
            full_score = 100 # 单项满分仍为100分制
            
            df_for_indicator = df_to_analyze
            # 对特定项目筛选适用的学生群体
            if indicator_key in ['1分钟仰卧起坐', '仰卧起坐(女)']:
                df_for_indicator = df_to_analyze[(df_to_analyze['学段'] == '小学') | ((df_to_analyze['学段'].isin(['初中', '高中'])) & (df_to_analyze['性别'] == '女'))].copy()
            elif indicator_key == '引体向上':
                df_for_indicator = df_to_analyze[(df_to_analyze['学段'].isin(['初中', '高中'])) & (df_to_analyze['性别'] == '男')].copy()
            elif indicator_key == '50米×8往返跑':
                df_for_indicator = df_to_analyze[pd.to_numeric(df_to_analyze['年级'], errors='coerce') == 5].copy()
            
            if df_for_indicator.empty:
                s2_log(f"项目 '{indicator_name}' 无适用学生，跳过总体统计。")
                continue

            s2_log(f"创建 {indicator_name} 总体_性别统计表...")
            overall_indicator_stats = _s2_calculate_group_indicator_stats(df_for_indicator, ['性别'], indicator_name, score_col_name, full_score, include_total_row=True)
            if not overall_indicator_stats.empty:
                overall_indicator_stats.to_excel(writer, sheet_name=sanitize_sheet_name(f'{indicator_name}总体_性别'), index=False)
        
        # --- 按选择的分组进行统计 ---
        if base_group_cols:
            s2_log(f"创建各 {group_by_selection} _总分_性别统计表...")
            group_cols_with_gender = base_group_cols + ['性别']
            group_total_stats = _s2_calculate_group_total_score_stats(df_to_analyze, group_cols_with_gender, score_thresholds, include_total_row=True)
            if not group_total_stats.empty:
                group_total_stats.to_excel(writer, sheet_name=f'各{group_by_selection}_总分_性别统计', index=False)

            for indicator_key, score_col_name in mapped_indicator_flat_cols.items():
                indicator_name = ALL_TESTS[indicator_key]['name']
                full_score = 100
                
                df_for_indicator = df_to_analyze
                if indicator_key in ['1分钟仰卧起坐', '仰卧起坐(女)']:
                    df_for_indicator = df_to_analyze[(df_to_analyze['学段'] == '小学') | ((df_to_analyze['学段'].isin(['初中', '高中'])) & (df_to_analyze['性别'] == '女'))].copy()
                elif indicator_key == '引体向上':
                    df_for_indicator = df_to_analyze[(df_to_analyze['学段'].isin(['初中', '高中'])) & (df_to_analyze['性别'] == '男')].copy()
                elif indicator_key == '50米×8往返跑':
                    df_for_indicator = df_to_analyze[pd.to_numeric(df_to_analyze['年级'], errors='coerce') == 5].copy()

                if df_for_indicator.empty:
                    s2_log(f"项目 '{indicator_name}' 无适用学生，跳过分组统计。")
                    continue
                
                s2_log(f"创建 {indicator_name} 各 {group_by_selection} _性别统计表...")
                group_indicator_stats = _s2_calculate_group_indicator_stats(df_for_indicator, group_cols_with_gender, indicator_name, score_col_name, full_score, include_total_row=True)
                if not group_indicator_stats.empty:
                    group_indicator_stats.to_excel(writer, sheet_name=sanitize_sheet_name(f'{indicator_name}各{group_by_selection}_性别'), index=False)

    # 准备返回数据
    output_buffer.seek(0)
    # 调用格式化函数
    formatted_buffer = _s2_format_excel_file(output_buffer, score_thresholds, mapped_indicator_flat_cols)
    return formatted_buffer.getvalue()

def _s2_process_and_analyze_data(input_file_buffer, file_name, group_by_selection, score_type, by_region, by_school, stage_selection):
    """主处理流程 (增加按学校拆分功能)"""
    global log_messages
    s2_log("", clear=True)
    try:
        s2_log(f"==================================================")
        s2_log(f"开始处理: {file_name}")
        df = pd.read_excel(input_file_buffer, header=[0, 1])
        s2_log("成功读取文件。")
        
        header_df = df.columns.to_frame(); header_df[0] = header_df[0].ffill()
        df.columns = pd.MultiIndex.from_frame(header_df)
        column_mapping = {}
        new_columns = []
        for c0, c1 in df.columns:
            key = (str(c0).strip(), str(c1).strip()); val = c0 if 'Unnamed' in str(c1) else f"{c0}_{c1}"
            column_mapping[key] = val; new_columns.append(val)
        df.columns = new_columns
        df = _s2_extract_grade_and_stage(df)
        if stage_selection != '所有学段':
            df = df[df['学段'] == stage_selection]
            if df.empty: s2_log(f"警告: 筛选后无数据，跳过。"); return None, None
        
        required_cols = ['姓名', '总分', '性别', '班级', '学段', '学校']
        if not all(col in df.columns for col in required_cols): raise ValueError(f"错误: 缺少关键列，必须包含: {required_cols}。")
        
        # 预先将所有分数转为数值
        all_score_cols = [col for col in df.columns if isinstance(col, str) and ('_分数' in col or col == '总分')]
        for col in all_score_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        mapped_indicator_flat_cols = {}
        for indicator_key, config in ALL_TESTS.items():
            if 'col_score' in config:
                parts = config['col_score'].rsplit('_', 1)
                if len(parts) == 2:
                    original_score_col_tuple = (parts[0].strip(), parts[1].strip())
                    flat_col_name = column_mapping.get(original_score_col_tuple)
                    if flat_col_name and flat_col_name in df.columns:
                        mapped_indicator_flat_cols[indicator_key] = flat_col_name
        
        if not mapped_indicator_flat_cols: s2_log("错误：未能匹配到任何分数-列。"); return None, None
        score_thresholds = TOTAL_SCORE_CATEGORIES_S2[score_type]
        output_filename_base = f"分析报告_{os.path.splitext(file_name)[0]}_{stage_selection}"
        
        # 决定处理模式
        if by_school:
            s2_log("模式: 按学校拆分报告")
            zip_buffer = io.BytesIO(); successful_files_in_zip = 0
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for school_name, school_df in df.groupby('学校'):
                    s2_log(f"\n--- 正在处理学校: {school_name} ---")
                    if school_df.empty: continue
                    result_bytes = _s2_perform_analysis_and_save(school_df, group_by_selection, score_thresholds, mapped_indicator_flat_cols)
                    if result_bytes:
                        zipf.writestr(f"{output_filename_base}_{sanitize_filename(school_name)}.xlsx", result_bytes); successful_files_in_zip += 1
            if successful_files_in_zip > 0:
                zip_buffer.seek(0); return zip_buffer.getvalue(), f"{output_filename_base}_按学校拆分.zip"
            else: s2_log("所有学校的处理均失败。"); return None, None
        elif by_region:
            s2_log("模式: 按区域拆分报告")
            df = _s2_add_region_column(df.copy())
            zip_buffer = io.BytesIO(); successful_files_in_zip = 0
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for region, region_df in df.groupby('区域'):
                    s2_log(f"\n--- 正在处理区域: {region} ---")
                    if region_df.empty: continue
                    result_bytes = _s2_perform_analysis_and_save(region_df, group_by_selection, score_thresholds, mapped_indicator_flat_cols)
                    if result_bytes:
                        zipf.writestr(f"{output_filename_base}_{region}.xlsx", result_bytes); successful_files_in_zip += 1
            if successful_files_in_zip > 0:
                zip_buffer.seek(0); return zip_buffer.getvalue(), f"{output_filename_base}_按区域.zip"
            else: s2_log("所有区域的处理均失败。"); return None, None
        else:
            s2_log("\n--- 正在对整体数据进行分析 ---")
            result_bytes = _s2_perform_analysis_and_save(df, group_by_selection, score_thresholds, mapped_indicator_flat_cols)
            return result_bytes, f"{output_filename_base}_整体.xlsx"
    except Exception as e:
        s2_log(f"处理文件 '{file_name}' 时发生严重错误: {e}"); s2_log(traceback.format_exc()); return None, None

def _s2_get_processed_dataframe(input_file_buffer, file_name):
    """
    读取并处理原始宽表Excel，转换为长表格式。
    """
    try:
        # 自动判断Excel文件类型 (.xls vs .xlsx)
        engine = 'xlrd' if file_name.lower().endswith('.xls') else 'openpyxl'
        df = pd.read_excel(input_file_buffer, engine=engine, header=[0, 1])
        
        # 修复复杂表头
        header_df = df.columns.to_frame()
        header_df[0] = header_df[0].ffill()
        df.columns = pd.MultiIndex.from_frame(header_df)
        
        new_columns = []
        for c0, c1 in df.columns:
            new_columns.append(c0 if 'Unnamed' in str(c1) else f"{c0}_{c1}")
        df.columns = new_columns
        
        # 删除完全无效的行
        df.dropna(subset=['姓名', '学籍号'], how='all', inplace=True)
        
        # 识别ID列和需要转换的数据列
        id_vars = [c for c in ['姓名', '学校', '班级', '学籍号', '总分'] if c in df.columns]
        value_vars = [c for c in df.columns if '_' in c]
        
        # 将宽表转换为长表 (melt)
        df_melted = df.melt(id_vars=id_vars, value_vars=value_vars, var_name='项目_类型', value_name='值')
        df_melted[['项目', '类型']] = df_melted['项目_类型'].str.split('_', expand=True, n=1)
        
        # 将“成绩”和“分数”重新透视回列 (pivot)
        df_final = df_melted.pivot_table(index=id_vars + ['项目'], columns='类型', values='值', aggfunc='first')
        df_final.rename_axis(None, axis=1, inplace=True)
        df_final = df_final.reset_index()

        # 处理可能存在的加分项
        bonus_cols = [c for c in df.columns if '加分' in c]
        if bonus_cols:
            df_bonus = df[id_vars + bonus_cols].melt(id_vars=id_vars, value_vars=bonus_cols, var_name='项目_加分', value_name='加分')
            df_bonus['项目'] = df_bonus['项目_加分'].str.replace('加分', '')
            df_final = pd.merge(df_final, df_bonus[['学籍号', '项目', '加分']], on=['学籍号', '项目'], how='left')
        else: 
            df_final['加分'] = ''
                
        # 数据清洗和格式化
        df_final.dropna(subset=['成绩'], inplace=True)
        if '学籍号' in df_final.columns:
            df_final['学籍号'] = df_final['学籍号'].astype(str).str.replace(r'\.0$', '', regex=True)
        
        for col in ['分数', '加分', '总分']:
            if col in df_final.columns:
                df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna('')
        
        df_final.fillna('', inplace=True)
        
        # 确保所有最终需要的列都存在
        final_cols_order = ['编号', '学籍号', '姓名', '学校', '班级', '项目', '成绩', '分数', '加分', '总分', '签名']
        for col in final_cols_order:
            if col not in df_final.columns:
                df_final[col] = ''
        
        # 排序并返回最终结果
        df_final.sort_values(by=['学校', '班级', '姓名'], inplace=True)
        
        return df_final[final_cols_order]
        
    except Exception as e:
        s2_log(f"读取或解析文件 '{file_name}' 失败。错误: {e}")
        st.error(f"读取或解析文件 '{file_name}' 失败。错误: {e}")
        st.exception(e)
        return None

def _s2_save_split_files(df_to_split, base_path_str, base_filename, split_by_class):
    """
    将处理好的DataFrame按学校或班级拆分并保存为多个Excel文件。
    """
    created_files_paths = []
    # 确保base_path是Path对象以便于操作
    base_path = Path(base_path_str)

    group_by_cols = ['学校', '班级'] if split_by_class else ['学校']
    
    if not all(col in df_to_split.columns for col in group_by_cols):
        st.error(f"拆分失败：源文件中缺少用于分组的列: {', '.join(group_by_cols)}")
        return []
            
    # 使用sort=False保留原始文件的顺序
    grouped = df_to_split.groupby(group_by_cols, sort=False, dropna=False)
    total_groups = grouped.ngroups
    
    if total_groups == 0:
        s2_log("在文件中没有找到可供分组的数据。")
        return []
    
    s2_log(f"检测到 {total_groups} 个分组组合，开始拆分文件...")
    
    for group_name, group_df in grouped:
        if isinstance(group_name, tuple):
            school_name = group_name[0]
            class_name = group_name[1] if len(group_name) > 1 else None
        else:
            school_name = group_name
            class_name = None
        
        # 创建目标文件夹
        school_folder = base_path / sanitize_filename(school_name)
        target_folder = school_folder / sanitize_filename(class_name) if class_name else school_folder
        try:
            target_folder.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            st.error(f"创建文件夹 '{target_folder}' 失败: {e}")
            continue

        # 将每个分组内的学生再次分割成小文件（每8人一个文件）
        unique_students = group_df['姓名'].unique()
        students_per_file = 8
        num_files = math.ceil(len(unique_students) / students_per_file)
        
        for i in range(num_files):
            student_chunk = unique_students[i * students_per_file : (i + 1) * students_per_file]
            chunk_df = group_df[group_df['姓名'].isin(student_chunk)].copy()
            
            # 重新排序并生成编号
            chunk_df.sort_values(by=['班级', '姓名'], inplace=True)
            chunk_df.reset_index(drop=True, inplace=True)
            chunk_df['编号'] = chunk_df.index + 1
            
            file_suffix = f"_Part_{i+1}" if num_files > 1 else ""
            output_filename = f"{base_filename}{file_suffix}.xlsx"
            output_path = target_folder / output_filename
            
            s2_log(f"    -> 正在保存文件: {output_path}")
            
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    chunk_df.to_excel(writer, sheet_name='学生成绩列表', index=False)
                    worksheet = writer.sheets['学生成绩列表']
                    _s2_apply_final_formatting(worksheet, chunk_df.columns.tolist())
                    _s2_merge_cells_by_key(worksheet, '姓名', chunk_df.columns.tolist())
                created_files_paths.append(output_path)
            except Exception as e:
                s2_log(f"    -> 保存文件 {output_filename} 失败: {e}")
                st.error(f"保存文件 {output_filename} 失败: {e}")

    return created_files_paths

def _s2_apply_final_formatting(worksheet, all_columns):
    """
    对给定的Excel工作表应用详细的格式化样式。
    """
    # 定义样式
    font_style = Font(name='微软雅黑', size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    project_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    wrap_text_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 定义精确的列宽
    exact_width_config = {
        '编号': 5.2, '学籍号': 10.3, '姓名': 10, '学校': 10.3, '班级': 6,
        '项目': 16, '成绩': 6, '分数': 6, '总分': 6, '加分': 6, '签名': 10
    }
    
    # 遍历所有单元格应用样式
    for row_idx in range(1, worksheet.max_row + 1):
        # 设置行高
        if row_idx > 1:
            worksheet.row_dimensions[row_idx].height = 13.5
        
        for col_idx in range(1, len(all_columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            col_name = all_columns[col_idx - 1]
            
            # 应用字体和边框
            cell.font = font_style
            cell.border = thin_border
            
            # 应用对齐方式 (数据行)
            if row_idx > 1:
                if col_name == '项目':
                    cell.alignment = project_align
                elif col_name in ['学籍号', '学校', '班级']:
                    cell.alignment = wrap_text_align
                else:
                    cell.alignment = center_align

    # 设置列宽
    for i, col_name in enumerate(all_columns, 1):
        col_letter = get_column_letter(i)
        if col_name in exact_width_config:
            # openpyxl的宽度单位与Excel略有不同，需要微调
            worksheet.column_dimensions[col_letter].width = exact_width_config[col_name]
            
    # 设置页面打印属性
    worksheet.page_setup.paper_size = worksheet.PAPERSIZE_A4
    worksheet.page_margins.left = worksheet.page_margins.right = 1.91 / 2.54  # 英寸转厘米
    worksheet.page_margins.top = worksheet.page_margins.bottom = 2.54 / 2.54
    worksheet.page_margins.header = worksheet.page_margins.footer = 1.27 / 2.54
    
def _s2_merge_cells_by_key(worksheet, key_column_name, all_columns):
    """
    根据给定的键列（如“姓名”），合并属于同一学生的指定列的单元格。
    """
    try: 
        # 找到关键列（用于判断是否为同一学生）的索引
        key_col_idx = all_columns.index(key_column_name) + 1
    except ValueError: 
        s2_log(f"警告: 在合并单元格时未找到关键列 '{key_column_name}'。")
        return # 如果找不到关键列，则无法执行合并

    # 定义需要被合并的列
    columns_to_merge_names = ['编号', '学籍号', '姓名', '学校', '班级', '总分', '签名']
    # 将列名转换为列索引
    cols_to_merge_indices = [all_columns.index(name) + 1 for name in columns_to_merge_names if name in all_columns]
    
    # 从第二行（第一个数据行）开始检查
    start_merge_row = 2
    
    # 遍历所有数据行
    # +2 是为了确保循环能处理到最后一行
    for row in range(3, worksheet.max_row + 2): 
        # 检查当前行是否与合并起始行的学生不同，或者是否已到达表格末尾
        if row > worksheet.max_row or \
           worksheet.cell(row=row, column=key_col_idx).value != worksheet.cell(row=start_merge_row, column=key_col_idx).value:
            
            # 如果合并区域包含多于一行，则执行合并
            if row - 1 > start_merge_row:
                for col_idx in cols_to_merge_indices:
                    try:
                        worksheet.merge_cells(start_row=start_merge_row, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                    except Exception as e:
                        s2_log(f"警告: 合并单元格区域 (行 {start_merge_row}-{row-1}, 列 {col_idx}) 失败: {e}")
            
            # 更新下一轮合并的起始行
            start_merge_row = row
    

def _s2_convert_excel_to_pdf(excel_file_path):
    """使用COM将单个Excel文件转换为PDF"""
    # 此功能在 Streamlit Cloud 等环境中不可用，仅本地Windows可用
    # Streamlit Cloud 用户可以考虑忽略PDF导出或使用其他库（如reportlab, fpdf）
    if not os.path.exists(excel_file_path):
        st.warning(f"PDF转换跳过：文件 {excel_file_path} 不存在。")
        return False
    
    excel_path = Path(excel_file_path).resolve()
    pdf_path = excel_path.with_suffix(".pdf")
    s2_log(f"  -> 正在将 '{excel_path.name}' 转换为 PDF...")
    try:
        import win32com.client # 尝试导入，如果失败会抛出ImportError
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(str(excel_path))
        worksheet = workbook.Worksheets[0] # 假设只有一个sheet
        worksheet.ExportAsFixedFormat(0, str(pdf_path)) # 0 = xlTypePDF
        s2_log(f"  -> 成功创建 PDF: {pdf_path.name}")
        return True
    except ImportError:
        st.error("PDF转换功能需要 `pywin32` 库，但未安装。请在Windows本地环境安装或忽略此功能。")
        return False
    except Exception as e:
        st.error(f"无法将 {excel_path.name} 转换为PDF。\n\n错误详情: {e}")
        st.exception(e)
        return False
    finally:
        if 'excel' in locals() and excel:
            excel.Quit()
    
def _s2_start_conversion(file_paths, to_pdf, split_files, split_by_class):
    success_count = 0
    fail_count = 0
    total_files = len(file_paths)
    zip_buffer = io.BytesIO()
    
    # 检查是否有权限写入PDF
    can_write_pdf = True
    try:
        import win32com.client
    except ImportError:
        st.error("PDF转换功能需要 `pywin32` 库，在当前环境不可用。")
        can_write_pdf = False
        if to_pdf:
            st.warning("PDF导出选项已被禁用，因为运行环境不支持。")
            to_pdf = False # 强制禁用PDF导出

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for i, file_path_str in enumerate(file_paths):
            file_path = Path(file_path_str)
            with st.status(f"处理中 ({i+1}/{total_files}): {file_path.name}...", expanded=True) as status:
                processed_df = _s2_get_processed_dataframe(file_path)
                if processed_df is None:
                    fail_count += 1
                    status.update(label=f"文件 {file_path.name} 处理失败!", state="error", expanded=True)
                    continue
                
                created_excel_files = []
                base_filename = f"打印模板_{file_path.stem}"
                
                if split_files:
                    # 目标文件夹为上传文件所在目录
                    created_excel_files = _s2_save_split_files(processed_df, file_path.parent, base_filename, split_by_class)
                else:
                    output_path = file_path.parent / f"{base_filename}.xlsx"
                    processed_df.sort_values(by=['班级', '姓名'], inplace=True)
                    processed_df['编号'] = processed_df.groupby('姓名', sort=False).ngroup() + 1
                    try:
                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            processed_df.to_excel(writer, sheet_name='学生成绩列表', index=False)
                            worksheet = writer.sheets['学生成绩列表']
                            _s2_apply_final_formatting(worksheet, processed_df.columns.tolist())
                            _s2_merge_cells_by_key(worksheet, '姓名', processed_df.columns.tolist())
                        created_excel_files.append(output_path)
                        s2_log(f"成功生成文件: {output_path.name}")
                    except Exception as e:
                        st.error(f"保存文件 {output_path.name} 失败: {e}")
                        fail_count += 1
                        continue # 继续处理下一个文件
                
                if created_excel_files:
                    if to_pdf and can_write_pdf:
                        s2_log(f"开始将 {len(created_excel_files)} 个Excel文件转换为PDF...")
                        pdfs_created = 0
                        for excel_file in created_excel_files:
                            if _s2_convert_excel_to_pdf(excel_file):
                                pdfs_created += 1
                        if pdfs_created == len(created_excel_files):
                            s2_log("所有PDF转换成功。")
                        elif pdfs_created > 0:
                            s2_log(f"部分PDF转换成功 ({pdfs_created}/{len(created_excel_files)})。", level='warning')
                        else:
                            s2_log("所有PDF转换均失败。", level='error')

                    # Add created Excel files to zip
                    for excel_file in created_excel_files:
                        zipf.write(excel_file, arcname=excel_file.name)
                    success_count +=1
                    status.update(label=f"文件 {file_path.name} 处理完成!", state="complete", expanded=False)
                else:
                    fail_count += 1
                    status.update(label=f"文件 {file_path.name} 未生成任何输出文件!", state="error", expanded=True)

    # Final check and download button
    if success_count > 0:
        st.session_state.s2_conversion_zip = zip_buffer.getvalue()
        st.success(f"批量转换完成！成功处理: {success_count} / {total_files}，失败: {fail_count}。")
        st.download_button(
            label="📥 **下载转换后的文件 (ZIP)**",
            data=st.session_state.s2_conversion_zip,
            file_name="转换后的成绩单文件.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary"
        )
    else:
        st.error(f"所有文件处理均失败。成功处理: {success_count} / {total_files}。请检查日志和文件路径。")
        st.session_state.s2_conversion_zip = None
    
def _s2_start_conversion_wrapper(file_paths, to_pdf, split_files, split_by_class):
    """包装器，用于处理Streamlit的文件上传和状态管理"""
    if not file_paths:
        st.error("请先选择要转换的Excel文件！")
        return
    
    # Clear previous download if new files are uploaded
    if 's2_conversion_zip' in st.session_state:
        st.session_state.s2_conversion_zip = None

    # Run the main conversion logic
    _s2_start_conversion(file_paths, to_pdf, split_files, split_by_class)
    
def display_logs(log_messages_list):
    if log_messages_list:
        st.expander("查看处理日志", expanded=False).text('\n'.join(log_messages_list))
    
def _s2_start_file_processing_tool(tool_name, process_func, *args):
    """通用启动器，用于处理文件操作工具"""
    try:
        with st.spinner(f"正在开始 '{tool_name}' 任务..."):
            process_func(*args)
        st.success(f"'{tool_name}' 任务完成！")
    except Exception as e:
        st.error(f"'{tool_name}' 任务执行失败: {e}")
        st.exception(e)
    finally:
        # Display any logs collected during the process
        display_logs(log_messages) # Ensure logs are displayed

def _worker_s2_fill_template(source_folder, template_file, output_folder):
    try:
        pd.set_option('display.float_format', lambda x: '%.0f' % x)
        header_mapping = {"姓名": "姓名(选填)", "性别": "性别(选填)", "年级": "年级(选填)", "班级": "班级(选填)", "学籍号": "学籍号(必填)", "跳绳个数": "跳绳"}
        
        source_files = [f for f in os.listdir(source_folder) if f.endswith(".xlsx") and not f.startswith('~$')]
        if not source_files:
            st.warning("在源文件夹中未找到Excel文件(.xlsx)")
            return

        total_files = len(source_files)
        success_count = 0
        for i, source_file_name in enumerate(source_files):
            source_file_path = Path(source_folder) / source_file_name
            with st.status(f"处理中 ({i+1}/{total_files}): {source_file_name}...", expanded=True) as status:
                try:
                    source_df = pd.read_excel(source_file_path)
                    target_wb = load_workbook(template_file)
                    target_ws = target_wb.active
                    target_headers = [cell.value for cell in target_ws[2] if cell.value] # Assuming header is in row 2

                    start_row = 3 # Data starts from row 3
                    for idx, source_row in source_df.iterrows():
                        for source_header, target_header in header_mapping.items():
                            if target_header in target_headers and source_header in source_row and pd.notna(source_row[source_header]):
                                col_idx = target_headers.index(target_header) + 1
                                original_value = source_row[source_header]
                                processed_value = original_value
                                if target_header == "年级(选填)": processed_value = _s3_replace_grade_value(original_value)
                                elif target_header == "班级(选填)": processed_value = _s3_replace_class_value(original_value)
                                elif target_header == "学籍号(必填)": processed_value = str(int(original_value)) if pd.notna(original_value) else ""
                                
                                target_ws.cell(row=start_row + idx, column=col_idx, value=processed_value)
                    
                    output_filename = os.path.join(output_folder, f"processed_{source_file_name}")
                    target_wb.save(output_filename)
                    success_count += 1
                    status.update(label=f"已处理: {source_file_name}", state="complete", expanded=False)
                except Exception as e:
                    st.error(f"处理文件 {source_file_name} 时出错: {e}")
                    st.exception(e)
                    fail_count += 1
                    status.update(label=f"文件 {source_file_name} 处理失败!", state="error", expanded=True)
        
        st.success(f"数据填充完成! 成功处理 {success_count} / {total_files} 个文件。")
    except Exception as e:
        st.error(f"数据填充过程中发生严重错误: {e}")
        st.exception(e)
    
def _s3_replace_grade_value(value):
    if pd.isna(value): return value
    value_str = str(value)
    if "小学2021级" in value_str: return "5"
    if "小学2023级" in value_str: return "3"
    if "初中2024级" in value_str: return "8"
    if "高中2024级" in value_str: return "11"
    return value
        
def _s3_replace_class_value(value):
    if pd.isna(value): return value
    value_str = str(value).replace("班", "").replace("2021级", "").replace("2023级", "").replace("2024级", "")
    return value_str if value_str else value
    
def _worker_s2_split_table(file_path, chunk_size):
    try:
        s2_log("正在读取文件，请稍候...")
        file_extension = file_path.suffix.lower()
        dtype_spec = {'学籍号': str, '班级': str}
        
        if file_extension == '.csv': df = pd.read_csv(file_path, dtype=dtype_spec)
        elif file_extension in ['.xls', '.xlsx']: df = pd.read_excel(file_path, dtype=dtype_spec)
        else:
            st.error(f"不支持的文件格式 '{file_extension}'。")
            return
        
        s2_log(f"文件读取成功，共 {len(df)} 行数据。")
        required_columns = ['学校', '年级', '班级', '学籍号', '序号']
        if not all(col in df.columns for col in required_columns):
            st.error(f"表格中缺少必需的列，需要: {', '.join(required_columns)}。")
            return

        dir_name = file_path.parent
        base_name = file_path.stem
        main_output_dir = dir_name / f"{base_name}_拆分结果"
        main_output_dir.mkdir(parents=True, exist_ok=True)

        grouped = df.groupby(['学校', '年级'], sort=False, dropna=False)
        total_groups = grouped.ngroups
        if total_groups == 0:
            s2_log("文件中没有可供分组的数据。")
            return
        
        s2_log(f"检测到 {total_groups} 个 学校-年级 组合，开始处理...")
        total_files_created = 0
        group_count = 0

        for (school, grade), group_df in grouped:
            group_count += 1
            s2_log(f"\n--- 正在处理 学校: {school}, 年级: {grade} ({len(group_df)}名学生) ---")
            
            grade_output_dir = main_output_dir / sanitize_filename(school) / sanitize_filename(grade)
            grade_output_dir.mkdir(parents=True, exist_ok=True)

            num_chunks = (len(group_df) - 1) // chunk_size + 1
            for i in range(num_chunks):
                chunk_df = group_df.iloc[i * chunk_size : (i + 1) * chunk_size].copy()
                chunk_df.sort_values(by=['班级', '学籍号'], ascending=True, inplace=True)
                chunk_df.reset_index(drop=True, inplace=True)
                chunk_df['序号'] = chunk_df.index + 1
                
                output_filename = f"{sanitize_filename(school)}_{sanitize_filename(grade)}_第{i+1}组{file_extension}"
                output_path = grade_output_dir / output_filename
                
                if file_extension == '.csv': chunk_df.to_csv(output_path, index=False, encoding='utf-8-sig')
                else: chunk_df.to_excel(output_path, index=False)
                
                s2_log(f"  已创建文件: {output_path.name}")
                total_files_created += 1

        s2_log(f"\n处理完成！共为 {total_groups} 个组合创建了 {total_files_created} 个文件。")
        s2_log(f"所有文件已保存在 '{main_output_dir}' 中。")
        st.success("表格拆分完成！")
    except Exception as e:
        st.error(f"处理过程中出现未知错误: {e}")
        st.exception(e)

def _worker_s2_archive_files(src_folder, files_per_folder):
    if not os.path.isdir(src_folder):
        st.error(f"错误：找不到文件夹 '{src_folder}'。")
        return

    files = [f for f in os.listdir(src_folder) if os.path.isfile(os.path.join(src_folder, f))]
    file_count = len(files)
    if file_count == 0:
        st.warning(f"在 '{src_folder}' 中没有找到需要整理的文件。")
        return
    
    s2_log(f"在 '{src_folder}' 中共找到 {file_count} 个文件。")
    s2_log(f"将按照每 {files_per_folder} 个文件一个文件夹进行整理...")

    moved_count = 0
    for i in range(0, file_count, files_per_folder):
        subfolder_name = f"subfolder_{i // files_per_folder + 1}"
        subfolder_path = os.path.join(src_folder, subfolder_name)
        os.makedirs(subfolder_path, exist_ok=True)
        
        files_to_move = files[i : i + files_per_folder]
        for file_name in files_to_move:
            try:
                shutil.move(os.path.join(src_folder, file_name), os.path.join(subfolder_path, file_name))
                moved_count += 1
            except OSError as e:
                s2_log(f"移动文件 '{file_name}' 时出错: {e}", level='warning')
    
    s2_log(f"\n✅ 所有 {moved_count} 个文件已整理完毕！")
    st.success("文件归档整理完成！")

def get_image_as_base64(path):
    """
    将本地图片文件转换为Base64字符串以便在HTML中显示。
    """
    try:
        with open(path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except FileNotFoundError:
        # 如果找不到logo文件，不报错，而是返回None
        return None

def _worker_s2_move_images(source_dir, target_dir, copy_instead):
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp', '.svg']
    os.makedirs(target_dir, exist_ok=True)
    moved_count = 0
    
    all_files = []
    for root, _, files in os.walk(source_dir):
        for file in files:
            if any(file.lower().endswith(ext) for ext in image_extensions):
                all_files.append(Path(root) / file)

    total_files = len(all_files)
    if total_files == 0:
        st.warning("在源文件夹中未找到任何图片文件。")
        return

    for i, source_path in enumerate(all_files):
        base_name = source_path.name
        target_path = Path(target_dir) / base_name
        
        counter = 1
        name, ext = os.path.splitext(base_name)
        while target_path.exists():
            target_path = Path(target_dir) / f"{name}_{counter}{ext}"
            counter += 1
            
        try:
            action = "复制" if copy_instead else "移动"
            if copy_instead: shutil.copy2(source_path, target_path)
            else: shutil.move(source_path, target_path)
            s2_log(f"{action}: {base_name} -> {target_path.name}")
            moved_count += 1
        except Exception as e:
            st.error(f"处理文件 {source_path} 时出错: {e}")
            st.exception(e)
    
    st.success(f"图片操作完成! 共处理 {moved_count} / {total_files} 个图片文件。")

def _worker_s2_group_and_filter(root_folder, start_suffix):
    GROUPING_RULES = [{'prefix': 'group1_id', 'start': 4, 'end': 15}, {'prefix': 'group2_year', 'start': 0, 'end': 4}]
    
    def _apply_retention_rules(directory, start_suffix_local):
        try:
            all_image_files = sorted([f for f in os.listdir(directory) if f.lower().endswith(('.jpg', '.jpeg', '.png')) and os.path.isfile(os.path.join(directory, f))])
        except FileNotFoundError:
            s2_log(f"文件夹 '{directory}' 不存在，跳过。")
            return
        
        if not all_image_files: return

        start_index = 0
        if start_suffix_local:
            if any(os.path.splitext(f)[0].endswith(start_suffix_local) for f in all_image_files):
                start_index = next((i for i, f in enumerate(all_image_files) if os.path.splitext(f)[0].endswith(start_suffix_local)), 0)
            else:
                s2_log(f"  未找到以 '{start_suffix_local}' 结尾的文件。将处理所有文件。")
        
        files_to_prescreen = all_image_files[:start_index]
        relevant_files = all_image_files[start_index:]
        num_to_keep = min(len(relevant_files), 50 if len(relevant_files) < 500 else (100 if len(relevant_files) < 1000 else 150))
        
        files_to_delete = files_to_prescreen + relevant_files[num_to_keep:]
        if not files_to_delete: return

        for filename in files_to_delete:
            try:
                os.remove(os.path.join(directory, filename))
            except OSError as e:
                s2_log(f"删除文件 {filename} 时出错: {e}", level='warning')

    def group_and_process_recursively(directory, rules, level=0):
        if level >= len(rules):
            _apply_retention_rules(directory, start_suffix)
            return

        rule = rules[level]
        prefix, start, end = rule['prefix'], rule['start'], rule['end']
        
        try:
            image_files = [f for f in os.listdir(directory) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')) and os.path.isfile(os.path.join(directory, f))]
        except FileNotFoundError: return
        
        if not image_files: return
        groups = defaultdict(list)
        for filename in image_files:
            name_without_ext = os.path.splitext(filename)[0]
            group_key = name_without_ext[start:end] if len(name_without_ext) >= end else 'short_name'
            groups[group_key].append(filename)
        
        if len(groups) <= 1:
            group_and_process_recursively(directory, rules, level + 1)
        else:
            s2_log(f"在 '{directory}' 中，按规则 '{prefix}' 发现 {len(groups)} 个组，开始移动文件...")
            next_level_folders = []
            for group_key, files_in_group in groups.items():
                new_subdir_path = os.path.join(directory, f"{prefix}_{group_key}")
                os.makedirs(new_subdir_path, exist_ok=True)
                next_level_folders.append(new_subdir_path)
                for filename in files_in_group:
                    if os.path.isfile(os.path.join(directory, filename)):
                        shutil.move(os.path.join(directory, filename), os.path.join(new_subdir_path, filename))
            
            for folder in next_level_folders:
                group_and_process_recursively(folder, rules, level + 1)
    
    all_dirs = [dirpath for dirpath, _, _ in os.walk(root_folder)]
    for i, dirpath in enumerate(all_dirs):
        is_generated_dir = any(dir_name.startswith(rule['prefix']) for dir_name in dirpath.replace(root_folder, '').split(os.sep) for rule in GROUPING_RULES)
        if is_generated_dir: continue
        s2_log(f"\n--- 扫描目录: {dirpath} ---")
        group_and_process_recursively(dirpath, GROUPING_RULES)

    s2_log("\n--- 所有文件夹处理完毕！ ---")
    st.success("分组与筛选操作完成！")


def _worker_s2_extract_names(folder_path, output_file, recursive):
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp', '.svg']
    image_names = []
    try:
        if recursive:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if any(file.lower().endswith(ext) for ext in image_extensions):
                        image_names.append(os.path.splitext(file)[0])
        else:
            for file in os.listdir(folder_path):
                if os.path.isfile(os.path.join(folder_path, file)) and any(file.lower().endswith(ext) for ext in image_extensions):
                    image_names.append(os.path.splitext(file)[0])
        
        unique_names = sorted(set(image_names))
        s2_log(f"找到 {len(unique_names)} 个唯一的图片名称:\n" + "-"*40)
        for name in unique_names: s2_log(name)
        
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                for name in unique_names: f.write(name + '\n')
            s2_log(f"\n结果已保存到: {output_file}")
        st.success("图片名称提取完成！")
    except Exception as e:
        st.error(f"提取名称过程中发生错误: {e}")
        st.exception(e)

def _worker_s2_cleanup_images(root_folder, whitelist_file):
    try:
        with open(whitelist_file, 'r', encoding='utf-8') as f:
            files_to_keep = {line.strip() for line in f if line.strip()}
        s2_log(f"成功加载了 {len(files_to_keep)} 个要保留的文件名。")

        all_files_to_scan = []
        image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp')
        for dirpath, _, filenames in os.walk(root_folder):
            for filename in filenames:
                if filename.lower().endswith(image_extensions):
                    all_files_to_scan.append(os.path.join(dirpath, filename))

        total_files = len(all_files_to_scan)
        deleted_files_count = 0
        s2_log("\n--- 开始扫描并删除图片 ---")
        for i, file_path in enumerate(all_files_to_scan):
            filename_without_ext = os.path.splitext(os.path.basename(file_path))[0]
            if filename_without_ext not in files_to_keep:
                try:
                    os.remove(file_path)
                    s2_log(f"  已删除: {file_path}")
                    deleted_files_count += 1
                except OSError as e:
                    s2_log(f"删除文件 {file_path} 时出错: {e}", level='warning')
        
        s2_log(f"\n图片删除完成，共删除了 {deleted_files_count} 个文件。")
        s2_log("\n--- 开始清理空文件夹 ---")
        
        deleted_folders_count = 0
        for dirpath, dirnames, filenames in os.walk(root_folder, topdown=False):
            if not dirnames and not filenames:
                try:
                    os.rmdir(dirpath)
                    s2_log(f"  已删除空文件夹: {dirpath}")
                    deleted_folders_count += 1
                except OSError as e:
                    s2_log(f"删除文件夹 {dirpath} 时出错: {e}", level='warning')
        s2_log(f"\n空文件夹清理完成，共删除了 {deleted_folders_count} 个文件夹。")
        st.success("白名单清理完成！")
    except FileNotFoundError:
        st.error(f"白名单文件 '{whitelist_file}' 未找到。")
    except Exception as e:
        st.error(f"白名单清理过程中发生错误: {e}")
        st.exception(e)
# =========================================================================
# =========================================================================
#  Streamlit UI 界面 (最终美化版)
# =========================================================================
# =========================================================================

def worker_generate_pdf(student_data_chunk):
    """
    工作函数，为一小块学生数据生成PDF。
    设计用于多进程并行处理。
    """
    results = []
    for student_data in student_data_chunk:
        student_name = student_data.get('姓名', f'未知_{student_data.name}')
        grade = student_data.get('年级', '未知')
        gender = student_data.get('性别', '')

        if grade == "未知" or not gender:
            results.append({'status': 'skipped', 'name': student_name, 'reason': '年级或性别信息缺失'})
            continue
        
        student_test_items = get_applicable_tests(grade, gender)
        if not student_test_items:
            results.append({'status': 'skipped', 'name': student_name, 'reason': '无法匹配测试项目'})
            continue

        try:
            # 为当前学生计算权重配置
            student_weight_config = WEIGHT_CONFIG['通用'].copy()
            grade_map = {
                '一年级': '小学一二年级', '二年级': '小学一二年级', '三年级': '小学三四年级', 
                '四年级': '小学三四年级', '五年级': '小学五六年级', '六年级': '小学五六年级',
            }
            matched_grade_key = next((grade_map[key] for key in grade_map if key in grade), None)
            if matched_grade_key:
                student_weight_config.update(WEIGHT_CONFIG[matched_grade_key])
            elif any(keyword in grade for keyword in ['初', '高', '大']):
                student_weight_config.update(WEIGHT_CONFIG['中学及以上'])

            plotly_fig = create_radar_chart(student_data, student_test_items, student_weight_config)
            if plotly_fig:
                img_bytes = plotly_fig.to_image(format="png", width=600, height=600, scale=2)
                pdf_data = generate_pdf_report(student_data, io.BytesIO(img_bytes), student_test_items)
                if pdf_data:
                    results.append({'status': 'success', 'name': student_name, 'data': pdf_data})
                else:
                    results.append({'status': 'skipped', 'name': student_name, 'reason': 'PDF生成返回空数据'})
            else:
                results.append({'status': 'skipped', 'name': student_name, 'reason': '无法生成雷达图'})
        except Exception as e:
            results.append({'status': 'skipped', 'name': student_name, 'reason': f'生成出错 - {e}'})
    return results

st.set_page_config(page_title="学生体质健康分析平台", layout="wide", initial_sidebar_state="expanded")

# --- 注入CSS样式 ---
st.markdown("""
<style>
    /* 全局变量和字体 */
    :root {
        --primary-color: #0068c9;
        --secondary-color: #f0f2f6; /* 页面背景 */
        --text-color: #31333F;
        --card-background-color: #ffffff;
        --border-color: #e0e0e0;
    }
    @media (prefers-color-scheme: dark) {
        :root {
            --primary-color: #1c83e1; --secondary-color: #0e1117;
            --text-color: #fafafa; --card-background-color: #161b22;
            --border-color: #30363d;
        }
    }
    body, .stApp { background-color: var(--secondary-color); color: var(--text-color); }
    h1, h2, h3, h4, h5, h6 { color: var(--text-color); }
    h1 { color: var(--primary-color); padding-bottom: 10px; border-bottom: 2px solid var(--primary-color); }
    .st-emotion-cache-1jicfl2, .st-emotion-cache-1wb2q5j {
        background-color: var(--card-background-color); border: 1px solid var(--border-color);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08); border-radius: 10px; padding: 20px !important;
    }
    .st-emotion-cache-1jicfl2 summary { font-size: 1.2rem; font-weight: 600; color: var(--primary-color); }
    .stButton > button {
        border-radius: 10px !important; font-weight: bold !important; transition: all 0.3s ease-in-out !important;
        border: 2px solid var(--primary-color) !important; color: var(--primary-color) !important;
        background-color: transparent !important;
    }
    .stButton > button:hover { transform: scale(1.02); box-shadow: 0 4px 8px rgba(0, 104, 201, 0.2); }
    .stButton > button[kind="primary"] { background-color: var(--primary-color) !important; color: white !important; }
    .stButton > button[kind="primary"]:hover { background-color: #0056b3 !important; border-color: #0056b3 !important; }
    [data-testid="stSidebar"] { background-color: var(--card-background-color); border-right: 1px solid var(--border-color); padding: 10px; }
    .stTextInput, .stMultiSelect, .stSelectbox { background-color: var(--secondary-color); }
    #MainMenu, footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# --- HTML生成函数 ---
def student_info_card(student_data, logo_base64):
    logo_html = f'<img src="data:image/png;base64,{logo_base64}" style="width: 60px; height: 60px; margin-right: 20px; vertical-align: middle;">' if logo_base64 else ""
    return f"""<div style="background: linear-gradient(135deg, var(--primary-color) 0%, #3f9efb 100%); padding: 25px; border-radius: 10px; color: white; display: flex; align-items: center; box-shadow: 0 4px 12px rgba(0,0,0,0.2);">{logo_html}<div><h2 style="color: white; margin: 0; padding: 0;">{student_data.get('姓名', 'N/A')}</h2><p style="color: white; margin: 5px 0 0 0; opacity: 0.9;"><strong>学校:</strong> {student_data.get('学校', 'N/A')} | <strong>班级:</strong> {student_data.get('班级', 'N/A')} | <strong>学籍号:</strong> {student_data.get('学籍号', 'N/A')}</p></div></div>"""

def create_timeline(suggestions):
    RATING_COLOR_MAP = {"优秀": "#28a745", "良好": "#17a2b8", "及格": "#ffc107", "不及格": "#dc3545", "未知": "#6c757d", "——": "#6c757d"}
    html_parts = [f"""<div style='margin-bottom: 30px; border: 1px solid var(--border-color); border-radius: 10px; padding: 15px; background-color: var(--secondary-color);'><h4 style='color: var(--primary-color);'>📌 训练总纲 (必读)</h4>"""]
    for item in TRAINING_SUGGESTIONS_DETAILED['训练总纲']: html_parts.append(f"<p style='color: var(--text-color);'><strong>{item['title']}:</strong> {item['content']}</p>")
    html_parts.append(f"</div><h4 style='color: var(--text-color);'><strong>▼ 针对性训练计划</strong></h4>")
    if not any(info['text'] for cat, info in suggestions.items()):
         html_parts.append("<p style='color: #28a745;'>太棒了！所有项目均表现良好或优秀，请继续保持！</p>")
    else:
        for cat, info in suggestions.items():
            if not info['text']: continue
            color = RATING_COLOR_MAP.get(info.get("rating", "未知"), "#6c757d")
            html_parts.append(f"""<div style='margin-bottom: 20px; position: relative; border-left: 3px solid var(--primary-color); padding-left: 30px;'><div style='width: 15px; height: 15px; background: var(--card-background-color); border: 3px solid var(--primary-color); border-radius: 50%; position: absolute; left: -9px; top: 5px;'></div><h5 style='margin-bottom: 5px; color: var(--text-color);'>{cat} (评级: <span style='color: {color}; font-weight: bold;'>{info['rating']}</span>)</h5><p style='margin: 0; color: var(--text-color); opacity: 0.9;'>{info['text'].replace('<br>', '<br/>').replace('<strong>', '<b>').replace('</strong>', '</b>')}</p></div>""")
    return "".join(html_parts)

# --- 侧边栏 ---
with st.sidebar:
    logo_path = ASSETS_DIR / "images" / "logo.png"
    st.image(
    str(logo_path) if logo_path.exists() else
    'https://www.google.com/images/branding/googlelogo/1x/googlelogo_color_272x92dp.png',
    width=100) 
    st.title("导航与信息")
    st.info("""**欢迎使用学生体质健康分析平台！**\n\n请在主界面选择您需要的功能。""")
    st.warning("部分工具箱功能需要在您自己的电脑上本地运行此应用。")

st.title("学生体质健康分析平台")
st.markdown("一站式完成 **个人报告生成** 与 **群体数据统计分析** 及其他工具。")

# --- 主功能标签页 ---
tab1, tab2 = st.tabs(["📊 **个人报告生成器**", "🛠️ **数据统计与工具箱**"])

with tab1:
    st.header("功能一：学生个人体质健康报告")
    st.markdown("上传包含学生体测成绩的 **原始Excel文件**，本工具将自动为每位学生生成详细的个人报告。")
    if 'processed_df' not in st.session_state: st.session_state.processed_df = None
    if 'pdf_to_download' not in st.session_state: st.session_state.pdf_to_download = None
    if 'zip_to_download' not in st.session_state: st.session_state.zip_to_download = None
    uploaded_file_s1 = st.file_uploader("📂 上传原始Excel文件 (用于生成个人报告)", type=['xlsx', 'xls'], key="uploader_s1")
    if uploaded_file_s1:
        st.session_state.pdf_to_download = None; st.session_state.zip_to_download = None; st.session_state.processed_df = None
        try:
            with st.spinner("正在读取和预处理文件..."):
                original_df = pd.read_excel(uploaded_file_s1, header=None)
                processed_df = process_complex_header(original_df)
                if processed_df is not None:
                    score_cols_100 = [col for col in processed_df.columns if isinstance(col, str) and '_分数' in col]
                    for col in score_cols_100:
                        if col in processed_df.columns:
                            processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce').fillna(0).astype(float) # Cast to float
                    for col in score_cols_100:
                        rating_col_name = col.replace('_分数', '_评级')
                        processed_df[rating_col_name] = processed_df[col].apply(get_rating)
                    if '班级' in processed_df.columns:
                        processed_df['年级'] = processed_df['班级'].apply(extract_grade_from_class)
                    processed_df = calculate_and_apply_weights(processed_df)
                    st.session_state.processed_df = processed_df
            st.success("✅ 文件预处理完成！现在可以生成报告了。")
        except Exception as e:
            st.error(f"文件处理失败: {e}"); st.exception(e); st.session_state.processed_df = None
    if st.session_state.get('processed_df') is not None:
        df_s1 = st.session_state.processed_df
        st.divider()
        
        required_cols = ['姓名', '班级', '性别', '学籍号', '学校']
        if not all(col in df_s1.columns for col in required_cols):
            st.error(f"错误：文件中必须包含以下几列: {', '.join(required_cols)}")
        else:
            with st.expander("🚀 **选项一：批量生成学生个人报告 (已优化速度)**", expanded=True):
                st.subheader("按学校筛选 (可选)")
                school_list = sorted([school for school in df_s1['学校'].unique() if pd.notna(school)])
                selected_schools = st.multiselect(
                    "选择一个或多个学校进行批量生成。如果留空，将生成所有学生的报告。",
                    options=school_list, key="school_multiselect"
                )
                if st.button("开始批量生成", key="batch_generate_s1", use_container_width=True, type="primary"):
                    if selected_schools:
                        df_to_process = df_s1[df_s1['学校'].isin(selected_schools)]
                        st.info(f"已选择 {len(selected_schools)} 所学校，将为 {len(df_to_process)} 名学生生成报告。")
                    else:
                        df_to_process = df_s1
                        st.info(f"未选择特定学校，将为文件中所有 {len(df_to_process)} 名学生生成报告。")
                    if df_to_process.empty:
                        st.warning("没有找到符合条件的学生，无法生成报告。")
                    else:
                        students_data = [row for index, row in df_to_process.iterrows()]
                        total_students = len(students_data)
                        cpu_cores = min(multiprocessing.cpu_count(), 8)
                        chunk_size = math.ceil(total_students / cpu_cores)
                        chunks = [students_data[i:i + chunk_size] for i in range(0, total_students, chunk_size)] if total_students > 0 else []
                        zip_buffer = io.BytesIO()
                        skipped_students = []; success_count = 0
                        with st.spinner(f"正在使用 {cpu_cores} 个CPU核心并行处理 {total_students} 名学生..."):
                            if chunks:
                                with multiprocessing.Pool(processes=cpu_cores) as pool:
                                    results_chunks = pool.map(worker_generate_pdf, chunks)
                                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                    for chunk in results_chunks:
                                        for result in chunk:
                                            if result['status'] == 'success':
                                                zip_file.writestr(f"{result['name']}_体质健康报告.pdf", result['data']); success_count += 1
                                            else: skipped_students.append(f"{result['name']} (原因: {result['reason']})")
                        st.success(f"🎉 批量处理完成！成功生成 {success_count} 份报告。")
                        if skipped_students: st.warning("以下学生因数据不完整或错误而被跳过："); st.json(skipped_students)
                        if success_count > 0: st.session_state.zip_to_download = zip_buffer.getvalue()
                        else: st.session_state.zip_to_download = None
            if st.session_state.get('zip_to_download'):
                zip_filename = "部分学生体质健康报告.zip" if st.session_state.get('school_multiselect') else "全体学生体质健康报告.zip"
                st.download_button(label="📥 **下载包含所选报告的ZIP文件**", data=st.session_state.zip_to_download, file_name=zip_filename, mime="application/zip", use_container_width=True)
            
            st.divider()
            st.header("🔍 **选项二：单个学生报告生成与预览**")
            col1, col2 = st.columns([1, 2])
            with col1:
                st.subheader("1. 选择学生")
                search_term = st.text_input("在此输入姓名或学籍号:", key="s1_search")
                filtered_df = df_s1[df_s1['姓名'].str.contains(search_term, case=False, na=False) | df_s1['学籍号'].astype(str).str.contains(search_term, case=False, na=False)] if search_term else df_s1
                student_list = filtered_df['姓名'].tolist()
                if not student_list:
                    st.warning("未找到匹配的学生。")
                    selected_student = None
                else:
                    selected_student = st.selectbox("从结果中选择学生:", student_list, key="student_selector", index=None, placeholder="请选择一名学生进行预览...")
            if selected_student:
                with col2:
                    st.subheader("2. 预览与下载")
                    student_data = df_s1[df_s1['姓名'] == selected_student].iloc[0]
                    grade = student_data.get('年级', '未知'); gender = student_data.get('性别', '')
                    logo_base64 = get_image_as_base64('logo.png')
                    st.markdown(student_info_card(student_data, logo_base64), unsafe_allow_html=True); st.divider()
                    student_test_items = get_applicable_tests(grade, gender)
                    if not student_test_items:
                        st.warning(f"无法为年级'{grade}'匹配到测试项目规则。")
                    else:
                        student_weight_config = WEIGHT_CONFIG['通用'].copy()
                        grade_map = {'一年级': '小学一二年级', '二年级': '小学一二年级', '三年级': '小学三四年级', '四年级': '小学三四年级', '五年级': '小学五六年级', '六年级': '小学五六年级'}
                        matched_grade_key = next((grade_map[key] for key in grade_map if key in grade), None)
                        if matched_grade_key: student_weight_config.update(WEIGHT_CONFIG[matched_grade_key])
                        elif any(keyword in grade for keyword in ['初', '高', '大']): student_weight_config.update(WEIGHT_CONFIG['中学及以上'])
                        plotly_fig = create_radar_chart(student_data, student_test_items, student_weight_config)
                        ptab1, ptab2, ptab3 = st.tabs(["📊 指标概览", "📈 指标详解", "🏃 训练计划"])
                        with ptab1:
                            m_col1, m_col2 = st.columns(2)
                            total_score_weighted = pd.to_numeric(student_data.get('总分'), errors='coerce')
                            m_col1.metric("体测总分 (加权后)", f"{total_score_weighted:.2f} 分")
                            m_col2.metric("综合评级", get_rating_30(total_score_weighted))
                            if plotly_fig: st.plotly_chart(plotly_fig, use_container_width=True)
                        with ptab2:
                            details_data = []
                            for item in student_test_items:
                                score_col = item.get('col_score'); rating_col = score_col.replace('_分数', '_评级') if score_col else None
                                name = item.get('name', 'N/A'); grade_val = student_data.get(item.get('col_grade'), 'N/A')
                                score_val = student_data.get(score_col, 'N/A'); rating_val = student_data.get(rating_col, '——')
                                formatted_score = f"{pd.to_numeric(score_val, errors='coerce'):.2f}" if pd.notna(pd.to_numeric(score_val, errors='coerce')) else 'N/A'
                                details_data.append([name, grade_val, formatted_score, rating_val])
                            details_df = pd.DataFrame(details_data, columns=['项目', '成绩', '加权得分', '评级'])
                            st.dataframe(details_df, use_container_width=True, hide_index=True)
                        with ptab3:
                            suggestions_web = {}
                            for item in student_test_items:
                                score_col = item.get('col_score')
                                if score_col:
                                    rating_col = score_col.replace('_分数', '_评级')
                                    rating = student_data.get(rating_col, "未知")
                                    text = TRAINING_SUGGESTIONS_DETAILED.get(item['name'], {}).get(gender, TRAINING_SUGGESTIONS_DETAILED.get(item['name'], {}).get("通用", {})).get(rating, "")
                                    if text: suggestions_web[item['name']] = {"rating": rating, "text": text}
                            st.markdown(create_timeline(suggestions_web), unsafe_allow_html=True)
                        if st.button("生成此学生的PDF报告", key=f"single_pdf_btn_{selected_student}", use_container_width=True):
                            with st.spinner("正在生成PDF文件..."):
                                if plotly_fig:
                                    img_bytes = plotly_fig.to_image(format="png", width=600, height=600, scale=2)
                                    pdf_data = generate_pdf_report(student_data, io.BytesIO(img_bytes), student_test_items)
                                    st.session_state.pdf_to_download = {"data": pdf_data, "name": f"{selected_student}_体质健康报告.pdf"}
                                else: st.warning("无法生成PDF，因为雷达图未能创建。")                      
                        if st.session_state.get('pdf_to_download'):
                            st.download_button(label="📥 点击下载已生成的PDF", data=st.session_state.pdf_to_download['data'], file_name=st.session_state.pdf_to_download['name'], mime="application/pdf", use_container_width=True)

# ========================== TAB 2: 数据统计与工具箱 ==========================
with tab2:
    st.header("功能二：数据统计与工具箱")
    st.markdown("提供批量数据分析、成绩单转换等多种实用工具。")

    with st.expander("📈 **群体数据统计分析**", expanded=True):
        st.markdown("上传 **一个或多个** 原始Excel文件，本工具将对数据进行深度统计，并生成格式化的分析报告。")
        
        with st.form("analysis_form"):
            col1, col2 = st.columns(2)
            with col1:
                group_by_choice = st.radio("选择数据分组方式", ['学校和班级', '学校', '班级', '无分组'], index=0, key="s2_group")
                score_type_choice = st.radio("选择总分分数段标准", list(TOTAL_SCORE_CATEGORIES_S2.keys()), index=0, key="s2_score_type")
            with col2:
                stage_choice = st.selectbox("选择分析学段", ['所有学段', '小学', '初中', '高中'], key="s2_stage")
                st.divider()
                st.write("**报告拆分选项 (可选):**")
                split_option = st.radio(
                    "选择报告生成方式",
                    ('合并报告', '按学校拆分报告', '按区域拆分报告'),
                    key="split_option",
                    horizontal=True,
                    label_visibility="collapsed"
                )

            uploaded_files_s2 = st.file_uploader("📂 上传一个或多个用于统计分析的Excel文件", type=['xlsx', 'xls'], accept_multiple_files=True, key="uploader_s2")
            submitted = st.form_submit_button("开始分析", use_container_width=True, type="primary")

            if submitted:
                st.session_state.s2_split_option_value = split_option

        if submitted and uploaded_files_s2:
            st.session_state.analysis_zip = None
            zip_buffer_s2 = io.BytesIO()
            successful_files_count = 0
            
            split_option_choice = st.session_state.s2_split_option_value
            by_school = (split_option_choice == '按学校拆分报告')
            by_region = (split_option_choice == '按区域拆分报告')

            with zipfile.ZipFile(zip_buffer_s2, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file in uploaded_files_s2:
                    with st.status(f"正在处理文件: {file.name}...", expanded=True) as status:
                        file_bytes = file.read()
                        analysis_result_bytes, output_filename = _s2_process_and_analyze_data(io.BytesIO(file_bytes), file.name, group_by_choice, score_type_choice, by_region, by_school)
                        
                        # 在status内部显示日志
                        log_container = st.expander("显示/隐藏处理日志", expanded=False)
                        with log_container:
                            st.text('\n'.join(log_messages))
                        
                        if analysis_result_bytes:
                            zipf.writestr(output_filename, analysis_result_bytes)
                            successful_files_count += 1
                            status.update(label=f"文件 {file.name} 处理完成!", state="complete", expanded=False)
                        else:
                            status.update(label=f"文件 {file.name} 处理失败! 详情见日志。", state="error", expanded=True)
            
            if successful_files_count > 0:
                st.success(f"🎉 分析完成！成功处理 {successful_files_count} / {len(uploaded_files_s2)} 个文件。")
                zip_buffer_s2.seek(0)
                st.session_state.analysis_zip = zip_buffer_s2.getvalue()
                st.session_state.successful_files_count = successful_files_count
            else:
                st.error(f"所有 {len(uploaded_files_s2)} 个文件均处理失败，请检查文件内容和格式，并查看日志获取详细信息。")
                st.session_state.analysis_zip = None
                st.session_state.successful_files_count = 0
        
        if 'analysis_zip' in st.session_state and st.session_state.analysis_zip:
            successful_files_count = st.session_state.get('successful_files_count', 0)
            split_option_submitted_value = st.session_state.get('s2_split_option_value', '合并报告')
            is_single_output_file = successful_files_count == 1 and split_option_submitted_value == '合并报告'
            
            if is_single_output_file:
                try:
                    with zipfile.ZipFile(io.BytesIO(st.session_state.analysis_zip)) as zf:
                        excel_filename = zf.namelist()[0]
                        excel_data = zf.read(excel_filename)
                        st.download_button(label="📥 **下载分析报告 (Excel)**", data=excel_data, file_name=excel_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                except Exception as e:
                    st.error(f"提取Excel文件时出错: {e}")
            else:
                zip_filename = f"数据统计分析报告_{split_option_submitted_value}.zip"
                st.download_button(label="📥 **下载包含所有分析报告的ZIP文件**", data=st.session_state.analysis_zip, file_name=zip_filename, mime="application/zip", use_container_width=True)

    with st.expander("📄 **成绩单转换与分发 (本地运行)**"):
        st.markdown("将体测中心导出的原始宽表格式Excel，转换为适合打印和分发的长表格式。")
        st.info("⚠️ **注意**：此功能涉及本地文件系统读写，**必须在您自己的电脑上运行此Streamlit App**，并提供本地文件的绝对路径。它无法在网页浏览器或在线部署环境中使用。")
        
        local_files_path = st.text_input("请输入包含Excel成绩模板文件的 **文件夹路径**：", key="s2_local_path_input")
        
        col1_conv, col2_conv = st.columns(2)
        with col1_conv:
            split_files_local = st.checkbox("按学校拆分文件到不同文件夹", key="s2_split_school_local")
            if split_files_local:
                split_by_class_local = st.checkbox("在学校文件夹内再按班级细分", key="s2_split_class_local")
            else:
                split_by_class_local = False
        with col2_conv:
            to_pdf_local = st.checkbox("同时导出为PDF (需要Windows和Excel环境)", key="s2_to_pdf_local")
            
        if st.button("开始转换本地文件夹中的文件", use_container_width=True):
            s2_log("", clear=True)
            if not local_files_path or not os.path.isdir(local_files_path):
                st.error("请输入一个有效的本地文件夹路径！")
            else:
                try:
                    excel_files = [os.path.join(local_files_path, f) for f in os.listdir(local_files_path) if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
                    if not excel_files:
                        st.warning("在指定路径下未找到任何Excel文件。")
                    else:
                        st.write(f"找到 {len(excel_files)} 个文件，开始处理...")
                        _s2_start_conversion(excel_files, to_pdf_local, split_files_local, split_by_class_local)
                except Exception as e:
                    st.error(f"处理本地文件时出错：{e}")
                    st.exception(e)
    
    with st.expander("🗂️ **更多文件处理工具 (本地运行)**"):
        st.warning("以下工具同样需要在本地运行Streamlit App，并提供本地文件/文件夹路径。")
        
        st.subheader("1. 数据填充到模板")
        s3_source_folder = st.text_input("源数据文件夹路径:", key="s3_source_folder_input")
        s3_template_file = st.text_input("目标模板文件路径:", key="s3_template_file_input")
        s3_output_folder = st.text_input("输出文件夹路径:", key="s3_output_folder_input")
        if st.button("开始填充", key="s3_run_button"):
            s2_log("", clear=True)
            if all([s3_source_folder, s3_template_file, s3_output_folder]) and \
               os.path.isdir(s3_source_folder) and os.path.isfile(s3_template_file) and os.path.isdir(s3_output_folder):
                _s2_start_file_processing_tool("数据填充", _worker_s2_fill_template, s3_source_folder, s3_template_file, s3_output_folder)
            else:
                st.error("请确保所有三个路径都已正确填写且有效。")

        st.subheader("2. 按学校年级拆分表格")
        s4_file_to_split_path = st.text_input("要拆分的表格文件路径:", key="s4_file_input")
        s4_chunk_size = st.number_input("每个小文件的数据行数:", min_value=1, value=30, key="s4_chunk_input")
        if st.button("开始拆分", key="s4_run_button"):
            s2_log("", clear=True)
            if s4_file_to_split_path and os.path.isfile(s4_file_to_split_path):
                _s2_start_file_processing_tool("表格拆分", _worker_s2_split_table, Path(s4_file_to_split_path), s4_chunk_size)
            else:
                st.error("请输入一个有效的文件路径。")

if __name__ == '__main__':
    multiprocessing.freeze_support()

























